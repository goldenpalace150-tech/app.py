import base64
import calendar
import hashlib
from datetime import datetime, timedelta
import io
import json
import sqlite3
import tempfile
import unicodedata
import re
import zoneinfo
import zipfile
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd
import requests
import streamlit as strlit

# ==========================================
# 0. RTL ARABIC TEXT & VISUAL CONFIG
# ==========================================
APP_VERSION = "BIO-ATTENDANCE-MONTHLY-REPORT-MOBILE-2026-09-05"

TEXT_CONFIG = {
    "page_title": "حضور وانصراف القصر الذهبي",
    "title_main": "✨ شركة القصر الذهبي ✨",
    "search_placeholder": "🔍 ابحث باسم الموظف أو رقم الكود...",
    "header_all": "👥 كافة موظفي الشركة النشطين ({})",
    "header_present": "🟢 المتواجدون / الحضور ({})",
    "header_late": "⏰ الموظفون المتأخرون ({})",
    "header_checkout": "🏁 المنصرفون ({})",
    "header_leave": "🏖️ الموظفون في إجازة ({})",
    "header_absent": "❌ الغيابات ({})",
    "err_api": "خطأ في الاتصال بواجهة BioTime: {}",
}

strlit.set_page_config(
    page_title=TEXT_CONFIG["page_title"],
    page_icon="📡",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ==========================================
# 1. CSS STYLING & ANIMATIONS
# ==========================================
strlit.markdown(
    """
    <style>
    header[data-testid="stHeader"] { display: none !important; }
    footer { display: none !important; }
    .stApp { direction: rtl; background-color: #f4f7f9; font-family: system-ui, -apple-system, sans-serif; }
    
    .block-container {
        padding-top: 15px !important;
        padding-bottom: 30px !important; 
        padding-left: 10px !important;
        padding-right: 10px !important;
        max-width: 100% !important;
    }

    .status-badge {
        display: flex;
        align-items: center;
        justify-content: center;
        background: linear-gradient(145deg, #ffffff, #f0f4f8);
        border: 1px solid #cbd5e1;
        padding: 8px 20px;
        border-radius: 30px;
        box-shadow: 0 4px 15px rgba(0,0,0,0.04);
        margin: 0 auto 15px auto;
        width: fit-content;
        gap: 12px;
    }
    
    .animated-dish {
        width: 34px;
        height: 34px;
        object-fit: contain;
        transform-origin: center center;
        animation: rotate-360 5s linear infinite;
    }
    
    @keyframes rotate-360 {
        0% { transform: rotate(0deg); }
        100% { transform: rotate(360deg); }
    }

    .status-indicator {
        display: flex;
        align-items: center;
        gap: 8px;
    }

    .blinking-dot {
        width: 10px;
        height: 10px;
        background-color: #22c55e;
        border-radius: 50%;
        display: inline-block;
        box-shadow: 0 0 0 0 rgba(34, 197, 94, 0.7);
        animation: pulse-green 1.5s infinite;
    }

    @keyframes pulse-green {
        0% { transform: scale(0.95); box-shadow: 0 0 0 0 rgba(34, 197, 94, 0.7); }
        70% { transform: scale(1); box-shadow: 0 0 0 6px rgba(34, 197, 94, 0); }
        100% { transform: scale(0.95); box-shadow: 0 0 0 0 rgba(34, 197, 94, 0); }
    }

    .online-text {
        font-size: 14px;
        font-weight: 800;
        color: #0f172a;
        letter-spacing: 0.5px;
    }

    div[data-testid="stColumn"] button {
        width: 100% !important;
        background: #ffffff !important;
        border-radius: 12px !important;
        padding: 12px 10px !important;
        text-align: center !important;
        box-shadow: 0 2px 6px rgba(0,0,0,0.04) !important;
        border: 1px solid #cbd5e1 !important;
        margin-bottom: 6px !important;
        transition: all 0.2s ease !important;
    }
    div[data-testid="stColumn"] button p {
        font-size: 13px !important;
        color: #1e293b !important;
        font-weight: 700 !important;
        margin: 0 !important;
        white-space: pre-line !important;
        line-height: 1.4 !important;
    }

    .responsive-grid-table {
        width: 100%;
        border-collapse: collapse;
        margin-top: 15px;
        font-size: 13px;
        background-color: #ffffff;
        border-radius: 10px;
        overflow: hidden;
        box-shadow: 0 4px 6px rgba(0,0,0,0.02);
    }
    .responsive-grid-table .table-main-title-header {
        background: #0f172a;
        color: #ffffff !important;
        text-align: center;
        font-size: 14px;
        padding: 12px;
    }
    .responsive-grid-table th {
        background-color: #f8fafc;
        color: #475569;
        padding: 10px;
        border-bottom: 1px solid #e2e8f0;
        text-align: center;
    }
    .responsive-grid-table td { 
        padding: 10px; border-bottom: 1px solid #f1f5f9; text-align: center; font-weight: 500; color: #1e293b;
    }
    .badge-present { background-color: #dcfce7; color: #166534; padding: 4px 8px; border-radius: 6px; font-size: 11px; }
    .badge-late { background-color: #fef3c7; color: #9a3412; padding: 4px 8px; border-radius: 6px; font-size: 11px; }
    .badge-leave { background-color: #e0f2fe; color: #0369a1; padding: 4px 8px; border-radius: 6px; font-size: 11px; }
    .badge-absent { background-color: #fee2e2; color: #991b1b; padding: 4px 8px; border-radius: 6px; font-size: 11px; }

    /* ===== Interface polish: cards, upload, progress, download, mobile ===== */
    .gp-section-title {
        font-size: 15px;
        font-weight: 800;
        color: #0f172a;
        margin: 8px 2px 10px 2px;
    }

    .gp-kpi-grid {
        display: grid;
        grid-template-columns: repeat(6, minmax(120px, 1fr));
        gap: 9px;
        margin: 0 0 14px 0;
    }

    .gp-kpi-card {
        background: #ffffff;
        border: 1px solid #e2e8f0;
        border-radius: 14px;
        padding: 11px 8px;
        text-align: center;
        box-shadow: 0 3px 12px rgba(15, 23, 42, 0.05);
        min-height: 68px;
        display: flex;
        flex-direction: column;
        justify-content: center;
    }

    .gp-kpi-icon { font-size: 16px; line-height: 1; margin-bottom: 5px; }
    .gp-kpi-value { font-size: 22px; line-height: 1.1; font-weight: 900; color: #0f172a; }
    .gp-kpi-label { margin-top: 4px; font-size: 11px; font-weight: 700; color: #64748b; }

    /* Clickable attendance summary cards (Streamlit buttons). */
    div[data-testid="stHorizontalBlock"] div[data-testid="stButton"] > button {
        min-height: 86px;
        border-radius: 14px !important;
        border: 1px solid #e2e8f0 !important;
        background: #ffffff !important;
        box-shadow: 0 3px 12px rgba(15, 23, 42, 0.05) !important;
        font-weight: 800 !important;
        line-height: 1.45 !important;
        white-space: pre-line !important;
    }

    .gp-file-note {
        background: #ffffff;
        border: 1px solid #e2e8f0;
        border-radius: 12px;
        padding: 9px 12px;
        margin-bottom: 7px;
        color: #475569;
        font-size: 12px;
        font-weight: 650;
        text-align: center;
    }

    div[data-testid="stFileUploader"] {
        background: #f8fafc;
        border: 1px solid #e2e8f0;
        border-radius: 14px;
        padding: 6px;
    }
    div[data-testid="stFileUploaderDropzone"] {
        border-radius: 11px !important;
        border-color: #cbd5e1 !important;
        background: #ffffff !important;
        padding-top: 10px !important;
        padding-bottom: 10px !important;
    }
    div[data-testid="stFileUploader"] small { display: none !important; }

    div[data-testid="stProgress"] { margin-top: 8px; margin-bottom: 8px; }

    .gp-download-ready {
        background: #ecfdf5;
        border: 1px solid #bbf7d0;
        color: #166534;
        border-radius: 12px;
        padding: 10px 12px;
        margin: 8px 0;
        text-align: center;
        font-weight: 800;
        font-size: 13px;
    }

    .gp-tech-note {
        color: #64748b;
        font-size: 11px;
        text-align: center;
        margin-top: 4px;
    }

    @media (max-width: 900px) {
        .gp-kpi-grid { grid-template-columns: repeat(3, minmax(100px, 1fr)); }
        .block-container { padding-left: 7px !important; padding-right: 7px !important; }
        .status-badge { padding: 7px 14px; margin-bottom: 10px; }
        .responsive-grid-table { font-size: 11px; display: block; overflow-x: auto; white-space: nowrap; }
    }

    @media (max-width: 560px) {
        .gp-kpi-grid { grid-template-columns: repeat(2, minmax(100px, 1fr)); gap: 7px; }
        .gp-kpi-card { min-height: 62px; padding: 9px 6px; }
        .gp-kpi-value { font-size: 19px; }
        .gp-kpi-label { font-size: 10px; }
        div[data-testid="stColumn"] { min-width: 100% !important; }
        div[data-testid="stColumn"] button { padding: 10px 8px !important; }
    }

    /* Full-screen loading veil used only while the app is waiting/processing. */
    .gp-loading-overlay {
        position: fixed; inset: 0; z-index: 999999;
        background: rgba(241, 245, 249, 0.78);
        backdrop-filter: blur(1.5px); -webkit-backdrop-filter: blur(1.5px);
        display: flex; align-items: center; justify-content: center; direction: rtl;
    }
    .gp-loading-box {
        min-width: 220px; background: rgba(255,255,255,0.96);
        border: 1px solid #cbd5e1; border-radius: 18px;
        box-shadow: 0 18px 55px rgba(15,23,42,0.16);
        padding: 22px 26px; text-align: center; color: #0f172a; font-weight: 800;
    }
    .gp-loading-spinner {
        width: 38px; height: 38px; margin: 0 auto 12px auto;
        border: 4px solid #dbeafe; border-top-color: #2563eb;
        border-radius: 50%; animation: gp-spin 0.8s linear infinite;
    }
    @keyframes gp-spin { to { transform: rotate(360deg); } }

    /* ===== App shell / responsive navigation ===== */
    .gp-app-hero {
        background: linear-gradient(135deg, #0f172a 0%, #1e3a8a 58%, #2563eb 100%);
        border-radius: 20px;
        padding: 18px 20px;
        margin: 0 0 12px 0;
        box-shadow: 0 14px 34px rgba(15, 23, 42, 0.16);
        color: #ffffff;
        text-align: center;
    }
    .gp-app-hero-title { font-size: 21px; font-weight: 900; line-height: 1.3; }
    .gp-app-hero-subtitle { font-size: 12px; opacity: 0.86; margin-top: 4px; font-weight: 650; }

    div[data-testid="stTabs"] div[data-baseweb="tab-list"] {
        background: rgba(255,255,255,0.92);
        border: 1px solid #dbe4ef;
        border-radius: 16px;
        padding: 5px;
        gap: 5px;
        box-shadow: 0 5px 18px rgba(15, 23, 42, 0.05);
        overflow-x: auto;
        scrollbar-width: none;
    }
    div[data-testid="stTabs"] div[data-baseweb="tab-list"]::-webkit-scrollbar { display: none; }
    div[data-testid="stTabs"] button[data-baseweb="tab"] {
        flex: 1 1 0;
        min-width: 130px;
        min-height: 46px;
        border-radius: 11px;
        font-weight: 800;
        color: #475569;
        white-space: nowrap;
    }
    div[data-testid="stTabs"] button[data-baseweb="tab"][aria-selected="true"] {
        background: #eff6ff;
        color: #1d4ed8;
    }
    div[data-testid="stTabs"] div[data-baseweb="tab-highlight"] { display: none; }

    .gp-report-panel {
        background: #ffffff;
        border: 1px solid #e2e8f0;
        border-radius: 16px;
        padding: 14px;
        margin: 8px 0 12px 0;
        box-shadow: 0 4px 18px rgba(15, 23, 42, 0.045);
    }
    .gp-report-title { font-size: 17px; font-weight: 900; color: #0f172a; text-align: center; }
    .gp-report-subtitle { font-size: 11px; color: #64748b; text-align: center; margin-top: 4px; }
    .gp-report-ready {
        background: linear-gradient(135deg, #ecfdf5, #f0fdf4);
        border: 1px solid #bbf7d0;
        color: #166534;
        border-radius: 14px;
        padding: 12px;
        font-weight: 850;
        text-align: center;
        margin: 10px 0;
    }

    @media (max-width: 700px) {
        .gp-app-hero { border-radius: 15px; padding: 14px 10px; margin-bottom: 9px; }
        .gp-app-hero-title { font-size: 17px; }
        .gp-app-hero-subtitle { font-size: 10px; }
        div[data-testid="stTabs"] div[data-baseweb="tab-list"] { padding: 4px; gap: 3px; }
        div[data-testid="stTabs"] button[data-baseweb="tab"] {
            min-width: 108px;
            min-height: 42px;
            padding-left: 7px !important;
            padding-right: 7px !important;
            font-size: 11px !important;
        }
        .gp-report-panel { padding: 10px; border-radius: 13px; }
    }


    /* ===== Compact responsive top controls ===== */
    .gp-top-controls {
        margin: 0 0 10px 0;
    }
    .gp-control-caption {
        color: #64748b;
        font-size: 10px;
        font-weight: 800;
        margin: 0 2px 5px 2px;
        min-height: 15px;
        line-height: 15px;
    }
    .gp-live-card,
    .gp-archive-card {
        min-height: 42px;
        width: 100%;
        border-radius: 12px;
        border: 1px solid #dbe4ef;
        background: #ffffff;
        box-shadow: 0 3px 12px rgba(15, 23, 42, 0.04);
        display: flex;
        align-items: center;
        justify-content: center;
        gap: 8px;
        padding: 7px 12px;
        box-sizing: border-box;
        white-space: nowrap;
    }
    .gp-live-card {
        color: #166534;
        background: linear-gradient(135deg, #f0fdf4, #ffffff);
        border-color: #bbf7d0;
    }
    .gp-archive-card {
        color: #475569;
        background: linear-gradient(135deg, #f8fafc, #ffffff);
        border-color: #cbd5e1;
    }
    .gp-live-dot {
        width: 9px;
        height: 9px;
        border-radius: 50%;
        background: #22c55e;
        box-shadow: 0 0 0 0 rgba(34, 197, 94, 0.55);
        animation: pulse-green 1.5s infinite;
        flex: 0 0 auto;
    }
    .gp-live-main { font-size: 13px; font-weight: 900; }
    .gp-live-sub { font-size: 10px; font-weight: 750; opacity: 0.72; }

    /* Make the native date control and refresh button look like one control family. */
    div[data-testid="stDateInput"] > div > div,
    div[data-testid="stDateInput"] input {
        min-height: 42px !important;
    }
    div[data-testid="stDateInput"] > div > div {
        border-radius: 12px !important;
        border-color: #dbe4ef !important;
        background: #ffffff !important;
        box-shadow: 0 3px 12px rgba(15, 23, 42, 0.04) !important;
    }
    .gp-refresh-wrap + div button,
    div[data-testid="stColumn"] button[kind="secondary"] {
        min-height: 42px !important;
    }

    @media (min-width: 701px) {
        .gp-top-controls-desktop-spacer { display: block; height: 1px; }
    }
    @media (max-width: 700px) {
        .gp-top-controls { margin-bottom: 7px; }
        .gp-control-caption { font-size: 9px; margin-bottom: 3px; }
        .gp-live-card, .gp-archive-card { min-height: 40px; border-radius: 11px; padding: 6px 10px; }
        .gp-live-main { font-size: 12px; }
        .gp-live-sub { font-size: 9px; }
        div[data-testid="stDateInput"] > div > div,
        div[data-testid="stDateInput"] input { min-height: 40px !important; }
    }

    </style>
""",
    unsafe_allow_html=True,
)

EXCLUDED_MANAGEMENT_CODES = ("40",)
# A lone punch at or after this hour is treated as an OUT punch.
SINGLE_PUNCH_OUT_HOUR = 14
# General Time Table boundaries visible in BioTime's Monthly Attendance Summary.
# When one side of a punch pair is missing, BioTime fills the missing side with
# the timetable boundary: 09:00 for missing IN, 19:00 for missing OUT.
SINGLE_PUNCH_SHIFT_START_HOUR = 9
SINGLE_PUNCH_SHIFT_END_HOUR = 19
SYRIA_TZ = zoneinfo.ZoneInfo("Asia/Damascus")

BASE_URL = strlit.secrets["biotime"]["base_url"].rstrip("/")
TOKEN_URL = strlit.secrets["biotime"]["token_url"]
EMAIL = strlit.secrets["biotime"]["email"]
PASSWORD = strlit.secrets["biotime"]["password"]
COMPANY = strlit.secrets["biotime"]["company"]

if "debug_logs" not in strlit.session_state:
  strlit.session_state["debug_logs"] = []
if "selected_view" not in strlit.session_state:
  strlit.session_state["selected_view"] = "present"


def clean_txt(raw_text):
  return (
      str(
          unicodedata.normalize("NFKC", str(raw_text))
          .replace("\u2066", "")
          .replace("\u2069", "")
          .strip()
      )
      if raw_text
      else ""
  )


def render_kpi_cards(items):
  """Render compact responsive summary cards without changing app logic."""
  cards = []
  for icon, label, value in items:
    cards.append(
        '<div class="gp-kpi-card">'
        f'<div class="gp-kpi-icon">{icon}</div>'
        f'<div class="gp-kpi-value">{value}</div>'
        f'<div class="gp-kpi-label">{label}</div>'
        '</div>'
    )
  strlit.markdown(
      '<div class="gp-kpi-grid">' + ''.join(cards) + '</div>',
      unsafe_allow_html=True,
  )



def _popup_dataframe(title, rows, search_key):
  """Open employee details only when a top summary card is clicked."""
  def render_content():
    if not rows:
      strlit.info("لا توجد سجلات في هذه الفئة.")
      return

    df = pd.DataFrame(rows)
    search_value = strlit.text_input(
        "",
        placeholder="🔍 ابحث باسم الموظف أو رقم الكود...",
        label_visibility="collapsed",
        key=f"popup_search_{search_key}",
    ).strip().casefold()

    if search_value:
      mask = df.astype(str).apply(
          lambda column: column.str.casefold().str.contains(
              search_value,
              regex=False,
              na=False,
          )
      ).any(axis=1)
      df = df[mask]

    strlit.dataframe(df, use_container_width=True, hide_index=True)

  if hasattr(strlit, "dialog"):
    dialog_renderer = strlit.dialog(title)(render_content)
    dialog_renderer()
  else:
    # Compatibility fallback for an older Streamlit runtime.
    with strlit.expander(title, expanded=True):
      render_content()


def render_clickable_attendance_cards(act, pre, lat, chk, lev, abs_s):
  """Render the six clean summary cards; details appear only in a dialog."""
  specs = [
      (
          "all",
          "👥",
          "الموظفون النشطون",
          len(act),
          [
              {"الكود": c, "الاسم": d["name"], "القسم": d["dept"], "الحالة": "نشط"}
              for c, d in act.items()
          ],
      ),
      (
          "present",
          "🟢",
          "الحضور الآن",
          len(pre),
          [
              {"الكود": c, "الاسم": n, "القسم": dpt, "الدخول": t, "الجهاز": d}
              for c, n, dpt, t, d in pre
          ],
      ),
      (
          "late",
          "⏰",
          "المتأخرون",
          len(lat),
          [
              {"الكود": c, "الاسم": n, "القسم": dpt, "الدخول": t, "الجهاز": d}
              for c, n, dpt, t, d in lat
          ],
      ),
      (
          "checkout",
          "🏁",
          "المنصرفون",
          len(chk),
          [
              {"الكود": c, "الاسم": n, "القسم": dpt, "الانصراف": t, "الجهاز": d}
              for c, n, dpt, t, d in chk
          ],
      ),
      (
          "leave",
          "🏖️",
          "الإجازات",
          len(lev),
          [
              {"الكود": c, "الاسم": n, "القسم": dpt, "نوع الإجازة": reason}
              for c, n, dpt, reason in lev
          ],
      ),
      (
          "absent",
          "❌",
          "الغيابات",
          len(abs_s),
          [
              {"الكود": c, "الاسم": n, "القسم": dpt, "الحالة": "غياب"}
              for c, n, dpt in abs_s
          ],
      ),
  ]

  cols = strlit.columns(6)
  for column, (key, icon, label, value, rows) in zip(cols, specs):
    with column:
      clicked = strlit.button(
          f"{icon}\n{value}\n{label}",
          key=f"summary_card_{key}",
          use_container_width=True,
      )
      if clicked:
        _popup_dataframe(f"{icon} {label} ({value})", rows, key)


def normalize_punch_to_minute(value):
  """Use the same minute precision displayed by BioTime reports.

  BioTime's Clock In / Clock Out report columns are minute-precision values.
  Using raw transaction seconds and flooring the final duration makes many
  locally calculated totals one minute shorter than BioTime.
  """
  if value is None:
    return None
  return value.replace(second=0, microsecond=0)


def calculate_single_punch_actual_wt(punch_time, is_out_punch, work_date=None):
  """Reproduce BioTime Monthly Attendance Summary for one missing punch.

  Missing IN  -> assume 09:00 and keep the real OUT, even when OUT is after
                 19:00 or shortly after midnight on the next calendar day.
  Missing OUT -> keep the real IN and assume 19:00 on the work date.

  This is intentionally NOT capped at ten hours. The user's BioTime report shows,
  for example, a missing IN with 21:32 OUT as 12:32 and a midnight OUT as 15:00.
  """
  punch_time = normalize_punch_to_minute(punch_time)
  if punch_time is None:
    return ""

  if work_date is None:
    # A punch shortly after midnight is normally the previous work day's OUT.
    if is_out_punch and punch_time.hour < 5:
      work_date = punch_time.date() - timedelta(days=1)
    else:
      work_date = punch_time.date()

  shift_start = datetime.combine(work_date, datetime.min.time()).replace(
      hour=SINGLE_PUNCH_SHIFT_START_HOUR
  )
  shift_end = datetime.combine(work_date, datetime.min.time()).replace(
      hour=SINGLE_PUNCH_SHIFT_END_HOUR
  )

  if is_out_punch:
    effective_in = shift_start
    effective_out = punch_time
  else:
    effective_in = punch_time
    effective_out = shift_end

  worked_seconds = int((effective_out - effective_in).total_seconds())
  if worked_seconds < 0:
    return ""

  total_minutes = worked_seconds // 60
  hours, minutes = divmod(total_minutes, 60)
  return f"{hours:02d}:{minutes:02d}"


def calculate_actual_wt_for_workday(work_date, clock_in, clock_out):
  """Return the single-punch value used by BioTime's monthly summary.

  For normal two-punch attendance the app uses Total WT instead. This function
  mainly exists for a missing IN or OUT and follows the General Time Table
  boundaries without the previous ten-hour cap.
  """
  if clock_in is None and clock_out is None:
    return ""

  if clock_in is None:
    return calculate_single_punch_actual_wt(
        clock_out,
        is_out_punch=True,
        work_date=work_date,
    )
  if clock_out is None:
    return calculate_single_punch_actual_wt(
        clock_in,
        is_out_punch=False,
        work_date=work_date,
    )

  # Kept for completeness; normal attendance exports Total WT below.
  return calculate_total_wt_for_workday(clock_in, clock_out)


def calculate_total_wt_for_workday(clock_in, clock_out):
  """BioTime Total Hrs: minute-precision interval between IN and OUT.

  Cross-midnight shifts are preserved. Example: 10:37 -> 00:29 next day = 13:52.
  """
  if clock_in is None or clock_out is None:
    return ""

  clock_in = normalize_punch_to_minute(clock_in)
  clock_out = normalize_punch_to_minute(clock_out)

  worked_seconds = int((clock_out - clock_in).total_seconds())
  if worked_seconds < 0:
    return ""

  total_minutes = worked_seconds // 60
  hours, minutes = divmod(total_minutes, 60)
  return f"{hours:02d}:{minutes:02d}"


def calculate_monthly_working_hours(work_date, clock_in, clock_out):
  """Final value written to Excel, matching BioTime Monthly Attendance Summary."""
  if clock_in is not None and clock_out is not None:
    return calculate_total_wt_for_workday(clock_in, clock_out)
  if clock_in is not None:
    return calculate_single_punch_actual_wt(
        clock_in,
        is_out_punch=False,
        work_date=work_date,
    )
  if clock_out is not None:
    return calculate_single_punch_actual_wt(
        clock_out,
        is_out_punch=True,
        work_date=work_date,
    )
  return ""


def calculate_odd_even_punch_total(punch_times):
  """Sum BioTime punches as sequential odd/even IN->OUT pairs.

  Punch 1 -> Punch 2, Punch 3 -> Punch 4, and so on. This is required
  when an employee has multiple attendance sessions in one work day; using only
  the first pair or a simple first-to-last interval gives the wrong duty time.

  Only complete pairs are summed here. A true one-punch day continues to use
  the existing single-punch Actual WT rule.
  """
  ordered = sorted(
      [value for value in punch_times if value is not None],
  )
  if len(ordered) < 2:
    return ""

  total_seconds = 0
  complete_pairs = 0
  for index in range(0, len(ordered) - 1, 2):
    pair_in = normalize_punch_to_minute(ordered[index])
    pair_out = normalize_punch_to_minute(ordered[index + 1])
    if pair_in is None or pair_out is None or pair_out < pair_in:
      continue
    total_seconds += int((pair_out - pair_in).total_seconds())
    complete_pairs += 1

  if complete_pairs == 0:
    return ""

  total_minutes = total_seconds // 60
  hours, minutes = divmod(total_minutes, 60)
  return f"{hours:02d}:{minutes:02d}"


def classify_transaction_punch(log, punch_time):
  """Return 'in', 'out', or 'unknown' from BioTime transaction punch state.

  BioTime transactions expose punch_state/punch_state_display. Respect those
  values first; only use the time of day when a device did not supply a usable
  state. This prevents multiple IN-only or OUT-only punches from being paired
  together into false 00:01/00:05/00:41 work durations.
  """
  display = clean_txt(log.get("punch_state_display", "")).casefold()
  raw_state = str(log.get("punch_state", "") or "").strip().casefold()

  display_compact = re.sub(r"[^a-z0-9]+", " ", display).strip()
  if display_compact in {
      "check in", "clock in", "in", "normal in", "overtime in", "ot in"
  }:
    return "in"
  if display_compact in {
      "check out", "clock out", "out", "normal out", "overtime out", "ot out"
  }:
    return "out"

  if raw_state in {"0", "in", "check in", "check-in", "clock in", "4"}:
    return "in"
  if raw_state in {"1", "out", "check out", "check-out", "clock out", "5"}:
    return "out"

  return "unknown"


def show_loading_overlay(message="جاري معالجة البيانات..."):
  """Fade the full app while a blocking operation is running."""
  placeholder = strlit.empty()
  safe_message = clean_txt(message)
  placeholder.markdown(
      f"""
      <div class="gp-loading-overlay">
        <div class="gp-loading-box">
          <div class="gp-loading-spinner"></div>
          <div>{safe_message}</div>
        </div>
      </div>
      """,
      unsafe_allow_html=True,
  )
  return placeholder


def hide_loading_overlay(placeholder):
  if placeholder is not None:
    placeholder.empty()


@strlit.cache_data(ttl=300)
def get_auth_token():
  try:
    payload = {
        "username": EMAIL,
        "email": EMAIL,
        "password": PASSWORD,
        "company": COMPANY,
    }
    res = requests.post(TOKEN_URL, json=payload, timeout=15)

    if res.status_code in (200, 201):
      return res.json().get("token")
    else:
      strlit.error(f"BioTime Server Error (Code {res.status_code}): {res.text}")
      return None

  except requests.exceptions.Timeout:
    strlit.error("Connection Failed: BioTime server timed out.")
    return None
  except Exception as e:
    strlit.error(f"Connection Failed: {str(e)}")
    return None



# ==========================================
# 2B. BIOTIME BACKUP (READ-ONLY)
# ==========================================
# Core endpoints documented by BioTime and already used by this app. Extra API
# routes are discovered read-only from API roots when BioTime exposes a route map.
BIOTIME_BACKUP_ENDPOINTS = (
    ("employees", ("/personnel/api/employees/",)),
    ("departments", ("/personnel/api/departments/",)),
    ("areas", ("/personnel/api/areas/",)),
    ("terminals", ("/iclock/api/terminals/", "/iclock/api/devices/")),
    ("transactions", ("/iclock/api/transactions/",)),
    ("leave", ("/att/api/leave/", "/iclock/api/leave/")),
)

BIOTIME_OPTIONAL_BACKUP_ENDPOINTS = (
    ("positions", ("/personnel/api/positions/",)),
    ("holidays", ("/att/api/holidays/",)),
    ("timetables", ("/att/api/timetables/",)),
    ("shifts", ("/att/api/shifts/",)),
    ("schedules", ("/att/api/schedules/", "/att/api/employee-schedules/")),
)


def _backup_json_default(value):
  if isinstance(value, (datetime,)):
    return value.isoformat()
  if hasattr(value, "isoformat"):
    try:
      return value.isoformat()
    except Exception:
      pass
  return str(value)


def _backup_extract_rows(payload):
  if isinstance(payload, list):
    return [item for item in payload if isinstance(item, dict)]
  if not isinstance(payload, dict):
    return []
  for key in ("data", "results", "items", "rows"):
    rows = payload.get(key)
    if isinstance(rows, list):
      return [item for item in rows if isinstance(item, dict)]
  # An object-info endpoint can legitimately return one record instead of a list.
  if payload and not any(key in payload for key in ("count", "next", "previous")):
    scalar_or_nested = any(not isinstance(value, str) or value for value in payload.values())
    if scalar_or_nested:
      return [payload]
  return []


def _backup_resolve_url(path_or_url):
  raw = str(path_or_url or "").strip()
  if raw.startswith("http://") or raw.startswith("https://"):
    return raw
  if not raw.startswith("/"):
    raw = "/" + raw
  return BASE_URL + raw


def _backup_fetch_pages(session, path_or_url, headers, page_size=5000, max_pages=10000):
  """Read a BioTime list endpoint completely without modifying server data."""
  url = _backup_resolve_url(path_or_url)
  params = {"page_size": page_size}
  pages = []
  rows = []
  seen_requests = set()
  status_code = None
  error = ""

  for page_number in range(1, max_pages + 1):
    request_key = (url, tuple(sorted((params or {}).items())))
    if request_key in seen_requests:
      break
    seen_requests.add(request_key)

    try:
      response = session.get(url, headers=headers, params=params, timeout=45)
    except requests.RequestException as exc:
      error = str(exc)
      break

    status_code = response.status_code
    if response.status_code != 200:
      error = f"HTTP {response.status_code}: {response.text[:500]}"
      break

    try:
      payload = response.json()
    except ValueError:
      error = "Response was not JSON"
      break

    pages.append(payload)
    page_rows = _backup_extract_rows(payload)
    rows.extend(page_rows)

    if isinstance(payload, dict):
      next_link = payload.get("next")
      if next_link:
        url = _backup_resolve_url(next_link)
        params = None
        continue

      count_value = payload.get("count")
      try:
        count_value = int(count_value)
      except (TypeError, ValueError):
        count_value = None

      if count_value is not None and len(rows) < count_value and page_rows:
        url = _backup_resolve_url(path_or_url)
        params = {"page_size": page_size, "page": page_number + 1}
        continue

    # No next link and no remaining count means the endpoint is complete.
    break

  return {
      "url": _backup_resolve_url(path_or_url),
      "status_code": status_code,
      "error": error,
      "pages": pages,
      "rows": rows,
  }


def _backup_try_endpoint_group(session, endpoint_paths, headers):
  attempts = []
  for endpoint in endpoint_paths:
    result = _backup_fetch_pages(session, endpoint, headers)
    attempts.append({
        "endpoint": endpoint,
        "status_code": result.get("status_code"),
        "error": result.get("error", ""),
    })
    if result.get("status_code") == 200:
      result["selected_endpoint"] = endpoint
      result["attempts"] = attempts
      return result
  return {
      "url": _backup_resolve_url(endpoint_paths[0]),
      "selected_endpoint": "",
      "status_code": attempts[-1]["status_code"] if attempts else None,
      "error": attempts[-1]["error"] if attempts else "No endpoint configured",
      "pages": [],
      "rows": [],
      "attempts": attempts,
  }


def _backup_safe_name(value):
  safe = re.sub(r"[^A-Za-z0-9_]+", "_", str(value or "").strip()).strip("_")
  return safe[:80] or "dataset"


def _backup_csv_bytes(rows):
  if not rows:
    return b""
  columns = []
  seen = set()
  for row in rows:
    for key in row.keys():
      key = str(key)
      if key not in seen:
        seen.add(key)
        columns.append(key)
  output = io.StringIO()
  import csv
  writer = csv.DictWriter(output, fieldnames=columns, extrasaction="ignore")
  writer.writeheader()
  for row in rows:
    flat = {}
    for column in columns:
      value = row.get(column)
      if isinstance(value, (dict, list, tuple)):
        value = json.dumps(value, ensure_ascii=False, default=_backup_json_default)
      elif value is None:
        value = ""
      flat[column] = value
    writer.writerow(flat)
  return output.getvalue().encode("utf-8-sig")


def _backup_sqlite_identifier(value, used):
  base = _backup_safe_name(value).lower()
  if base and base[0].isdigit():
    base = "c_" + base
  candidate = base or "column"
  suffix = 2
  while candidate in used:
    candidate = f"{base}_{suffix}"
    suffix += 1
  used.add(candidate)
  return candidate


def _backup_write_sqlite(db_path, datasets, manifest):
  conn = sqlite3.connect(db_path)
  try:
    conn.execute(
        "CREATE TABLE backup_manifest (key TEXT PRIMARY KEY, value TEXT NOT NULL)"
    )
    for key, value in manifest.items():
      conn.execute(
          "INSERT OR REPLACE INTO backup_manifest(key, value) VALUES (?, ?)",
          (str(key), json.dumps(value, ensure_ascii=False, default=_backup_json_default)),
      )

    used_tables = set()
    for dataset_name, dataset in datasets.items():
      rows = dataset.get("rows", [])
      table_name = _backup_sqlite_identifier(dataset_name, used_tables)
      if not rows:
        conn.execute(
            f'CREATE TABLE "{table_name}" (_record_json TEXT NOT NULL)'
        )
        continue

      original_columns = []
      seen_original = set()
      for row in rows:
        for key in row.keys():
          key = str(key)
          if key not in seen_original:
            seen_original.add(key)
            original_columns.append(key)

      used_columns = {"_record_json"}
      column_map = {}
      for original in original_columns:
        column_map[original] = _backup_sqlite_identifier(original, used_columns)

      definitions = [f'"{column_map[column]}" TEXT' for column in original_columns]
      definitions.append('"_record_json" TEXT NOT NULL')
      conn.execute(f'CREATE TABLE "{table_name}" ({", ".join(definitions)})')

      insert_columns = [column_map[column] for column in original_columns] + ["_record_json"]
      placeholders = ",".join("?" for _ in insert_columns)
      quoted_columns = ",".join(f'"{column}"' for column in insert_columns)
      sql = f'INSERT INTO "{table_name}" ({quoted_columns}) VALUES ({placeholders})'

      values_to_insert = []
      for row in rows:
        values = []
        for original in original_columns:
          value = row.get(original)
          if isinstance(value, (dict, list, tuple)):
            value = json.dumps(value, ensure_ascii=False, default=_backup_json_default)
          elif value is None:
            value = None
          else:
            value = str(value)
          values.append(value)
        values.append(json.dumps(row, ensure_ascii=False, default=_backup_json_default))
        values_to_insert.append(values)
      conn.executemany(sql, values_to_insert)
    conn.commit()
  finally:
    conn.close()


def _backup_discover_routes(session, headers):
  """Read API root maps when BioTime exposes them; never performs POST/PATCH/DELETE."""
  catalogs = {}
  discovered = {}
  base_host = re.sub(r"^https?://", "", BASE_URL).split("/", 1)[0]

  for root in ("/personnel/api/", "/iclock/api/", "/att/api/"):
    url = _backup_resolve_url(root)
    try:
      response = session.get(url, headers=headers, timeout=20)
      status = response.status_code
      payload = response.json() if status == 200 else {"error": response.text[:500]}
    except Exception as exc:
      status = None
      payload = {"error": str(exc)}
    catalogs[root] = {"status_code": status, "payload": payload}

    if status != 200 or not isinstance(payload, dict):
      continue
    for key, value in payload.items():
      if not isinstance(value, str):
        continue
      candidate = value.strip()
      if not (candidate.startswith("http://") or candidate.startswith("https://") or candidate.startswith("/")):
        continue
      absolute = _backup_resolve_url(candidate)
      candidate_host = re.sub(r"^https?://", "", absolute).split("/", 1)[0]
      if candidate_host != base_host or "/api/" not in absolute:
        continue
      dataset_key = f"discovered_{_backup_safe_name(root)}_{_backup_safe_name(key)}"
      discovered[dataset_key] = absolute
  return catalogs, discovered


def _backup_collect_media_urls(value, found, key_hint=""):
  if isinstance(value, dict):
    for key, child in value.items():
      _backup_collect_media_urls(child, found, str(key))
    return
  if isinstance(value, list):
    for child in value:
      _backup_collect_media_urls(child, found, key_hint)
    return
  if not isinstance(value, str):
    return

  key_lower = str(key_hint or "").lower()
  if not any(marker in key_lower for marker in ("photo", "image", "avatar", "picture")):
    return
  raw = value.strip()
  if not raw or raw.startswith("data:"):
    return
  if raw.startswith("http://") or raw.startswith("https://") or raw.startswith("/media/"):
    found.add(_backup_resolve_url(raw))


def _backup_download_media(session, headers, datasets, zip_handle, manifest):
  urls = set()
  for dataset in datasets.values():
    for row in dataset.get("rows", []):
      _backup_collect_media_urls(row, urls)

  base_host = re.sub(r"^https?://", "", BASE_URL).split("/", 1)[0]
  downloaded = 0
  failed = []
  for index, url in enumerate(sorted(urls), start=1):
    host = re.sub(r"^https?://", "", url).split("/", 1)[0]
    if host != base_host:
      failed.append({"url": url, "reason": "external host skipped"})
      continue
    try:
      response = session.get(url, headers=headers, timeout=30)
      if response.status_code != 200:
        failed.append({"url": url, "reason": f"HTTP {response.status_code}"})
        continue
      content_type = response.headers.get("Content-Type", "application/octet-stream")
      extension = ".bin"
      if "jpeg" in content_type or "jpg" in content_type:
        extension = ".jpg"
      elif "png" in content_type:
        extension = ".png"
      elif "webp" in content_type:
        extension = ".webp"
      elif "gif" in content_type:
        extension = ".gif"
      zip_handle.writestr(f"media/media_{index:05d}{extension}", response.content)
      downloaded += 1
    except requests.RequestException as exc:
      failed.append({"url": url, "reason": str(exc)})

  manifest["media"] = {
      "urls_found": len(urls),
      "downloaded": downloaded,
      "failed": failed,
  }


def build_biotime_backup(progress_callback=None):
  """Create a read-only BioTime snapshot: raw JSON + CSV + SQLite + manifest."""
  token = get_auth_token()
  if not token:
    raise RuntimeError("تعذر المصادقة مع BioTime")

  headers = {
      "Authorization": f"Token {token}",
      "Accept": "application/json",
  }
  session = requests.Session()
  datasets = {}
  manifest = {
      "backup_format_version": 1,
      "created_at": datetime.now(SYRIA_TZ).isoformat(),
      "source_base_url": BASE_URL,
      "company": COMPANY,
      "app_version": APP_VERSION,
      "read_only": True,
      "datasets": {},
      "restore_note": (
          "Raw JSON is the authoritative copy. BioTime Cloud restore capability depends "
          "on which write/import APIs ZKTeco exposes to this tenant; this ZIP is not a "
          "server/database image."
      ),
  }

  endpoint_groups = list(BIOTIME_BACKUP_ENDPOINTS) + list(BIOTIME_OPTIONAL_BACKUP_ENDPOINTS)
  total_steps = len(endpoint_groups) + 2
  completed_steps = 0

  for dataset_name, endpoint_paths in endpoint_groups:
    if progress_callback:
      progress_callback(completed_steps / total_steps, f"جاري نسخ {dataset_name}...")
    result = _backup_try_endpoint_group(session, endpoint_paths, headers)
    datasets[dataset_name] = result
    manifest["datasets"][dataset_name] = {
        "endpoint": result.get("selected_endpoint", ""),
        "status_code": result.get("status_code"),
        "rows": len(result.get("rows", [])),
        "pages": len(result.get("pages", [])),
        "error": result.get("error", ""),
        "attempts": result.get("attempts", []),
    }
    completed_steps += 1

  if progress_callback:
    progress_callback(completed_steps / total_steps, "جاري اكتشاف بيانات API إضافية...")
  api_catalogs, discovered_routes = _backup_discover_routes(session, headers)
  manifest["api_catalogs"] = {
      root: {"status_code": info.get("status_code")}
      for root, info in api_catalogs.items()
  }

  known_urls = {
      _backup_resolve_url(path)
      for _name, paths in endpoint_groups
      for path in paths
  }
  # Back up additional list endpoints advertised by BioTime, up to a generous
  # limit. Failures are recorded rather than aborting the whole snapshot.
  for dataset_name, endpoint_url in list(discovered_routes.items())[:100]:
    if endpoint_url in known_urls:
      continue
    result = _backup_fetch_pages(session, endpoint_url, headers)
    if result.get("status_code") == 200:
      datasets[dataset_name] = result
      manifest["datasets"][dataset_name] = {
          "endpoint": endpoint_url,
          "status_code": result.get("status_code"),
          "rows": len(result.get("rows", [])),
          "pages": len(result.get("pages", [])),
          "error": result.get("error", ""),
      }

  completed_steps += 1
  if progress_callback:
    progress_callback(completed_steps / total_steps, "جاري إنشاء JSON / CSV / SQLite...")

  output = io.BytesIO()
  with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=6) as zout:
    for dataset_name, dataset in datasets.items():
      safe_name = _backup_safe_name(dataset_name)
      raw_payload = {
          "source_url": dataset.get("url"),
          "status_code": dataset.get("status_code"),
          "error": dataset.get("error", ""),
          "pages": dataset.get("pages", []),
      }
      zout.writestr(
          f"raw/{safe_name}.json",
          json.dumps(raw_payload, ensure_ascii=False, indent=2, default=_backup_json_default),
      )
      csv_bytes = _backup_csv_bytes(dataset.get("rows", []))
      zout.writestr(f"csv/{safe_name}.csv", csv_bytes)

    zout.writestr(
        "raw/api_catalogs.json",
        json.dumps(api_catalogs, ensure_ascii=False, indent=2, default=_backup_json_default),
    )

    with tempfile.NamedTemporaryFile(suffix=".sqlite", delete=False) as temp_db:
      temp_db_path = temp_db.name
    try:
      _backup_write_sqlite(temp_db_path, datasets, manifest)
      with open(temp_db_path, "rb") as db_file:
        zout.writestr("sqlite/biotime_backup.sqlite", db_file.read())
    finally:
      try:
        import os
        os.remove(temp_db_path)
      except OSError:
        pass

    _backup_download_media(session, headers, datasets, zout, manifest)

    restore_readme = "BioTime Cloud backup package\n\n"
    restore_readme += "1. raw/*.json contains the authoritative API responses page-by-page.\n"
    restore_readme += "2. csv/*.csv is provided for inspection and spreadsheet recovery.\n"
    restore_readme += "3. sqlite/biotime_backup.sqlite contains a queryable snapshot.\n"
    restore_readme += "4. media/ contains same-host photos/images that were exposed as downloadable URLs.\n"
    restore_readme += "5. Restore into BioTime must be tested on a safe tenant first. Master data (departments/areas) should be restored before employees, followed by schedules/configuration and then transactions.\n"
    restore_readme += "6. This package cannot contain server-only database tables, biometric templates, or configuration that BioTime Cloud does not expose to the account/API.\n"
    zout.writestr("RESTORE_README.txt", restore_readme)
    zout.writestr(
        "manifest.json",
        json.dumps(manifest, ensure_ascii=False, indent=2, default=_backup_json_default),
    )

  output.seek(0)
  if progress_callback:
    progress_callback(1.0, "اكتملت النسخة الاحتياطية.")
  return output.getvalue(), manifest


def load_attendance_data_from_api(selected_date_str, selected_date_obj, is_today):
  token = get_auth_token()
  if not token:
    raise Exception(
        "تعذر المصادقة مع السيرفر. يرجى مراجعة تفاصيل الخطأ بالأعلى."
    )
  headers = {
      "Authorization": f"Token {token}",
      "Content-Type": "application/json",
  }

  devices = []
  try:
    for endpoint in ["/iclock/api/terminals/", "/iclock/api/devices/"]:
      dev_res = requests.get(
          f"{BASE_URL}{endpoint}", headers=headers, timeout=10
      )
      if dev_res.status_code == 200:
        d_data = dev_res.json()
        devices = (
            d_data.get("data", d_data)
            if isinstance(d_data, (dict, list))
            else []
        )
        break
  except Exception:
    pass

  terminal_map = {
      str(d.get("sn", "")): (
          d.get("alias") or d.get("terminal_name") or str(d.get("sn", ""))
      )
      for d in devices
      if d.get("sn")
  }

  all_employees = []
  try:
    emp_res = requests.get(
        f"{BASE_URL}/personnel/api/employees/?page_size=1000",
        headers=headers,
        timeout=15,
    )
    if emp_res.status_code == 200:
      all_employees = emp_res.json().get("data", [])
  except Exception:
    pass

  active_employees = {}
  for emp in all_employees:
    raw_code = str(emp.get("emp_code", "")).strip()
    cleaned_code = str(int(raw_code)) if raw_code.isdigit() else raw_code

    is_active = (
        str(emp.get("is_active", True)).lower() in ("true", "1", "yes")
    )
    emp_status = str(emp.get("status", "0")).upper()
    enable_att = (
        str(emp.get("enable_attendance", True)).lower() in ("true", "1", "yes")
    )

    if not is_active or emp_status in ("1", "2", "D") or not enable_att:
      continue

    if cleaned_code and cleaned_code not in EXCLUDED_MANAGEMENT_CODES:
      f_name = str(emp.get("first_name", "")).strip()
      l_name = str(emp.get("last_name", "")).strip()
      if f_name.lower() == "none":
        f_name = ""
      if l_name.lower() == "none":
        l_name = ""
      full_name = f"{f_name} {l_name}".strip()

      dept_data = emp.get("department", {})
      dept_name = (
          dept_data.get("dept_name")
          if isinstance(dept_data, dict)
          else str(emp.get("department", ""))
      )
      if not dept_name or dept_name.lower() == "none":
        dept_name = "غير محدد"

      active_employees[cleaned_code] = {
          "name": clean_txt(full_name if full_name else f"موظف {cleaned_code}"),
          "dept": clean_txt(dept_name),
      }

  leave_records = []
  try:
    leave_res = requests.get(
        f"{BASE_URL}/att/api/leave/?page_size=1000", headers=headers, timeout=10
    )
    if leave_res.status_code == 200:
      l_data = leave_res.json()
      leave_records = (
          l_data.get("data", l_data) if isinstance(l_data, (dict, list)) else []
      )
  except Exception:
    try:
      leave_res = requests.get(
          f"{BASE_URL}/iclock/api/leave/?page_size=1000",
          headers=headers,
          timeout=10,
      )
      if leave_res.status_code == 200:
        l_data = leave_res.json()
        leave_records = (
            l_data.get("data", l_data)
            if isinstance(l_data, (dict, list))
            else []
        )
    except Exception:
      pass

  on_leave_employees = {}
  for leave in leave_records:
    raw_code = str(
        leave.get("emp_code") or leave.get("employee_code") or ""
    ).strip()
    cleaned_code = str(int(raw_code)) if raw_code.isdigit() else raw_code
    start_t = (
        leave.get("start_time")
        or leave.get("start_date")
        or leave.get("start_datetime")
    )
    end_t = (
        leave.get("end_time")
        or leave.get("end_date")
        or leave.get("end_datetime")
    )

    leave_name = "إجازة"
    if "leave_type" in leave:
      if isinstance(leave["leave_type"], dict):
        leave_name = leave["leave_type"].get("leave_name", "إجازة")
      else:
        leave_name = str(leave["leave_type"])
    elif "leave_name" in leave:
      leave_name = leave["leave_name"]

    if cleaned_code and start_t and end_t:
      try:
        s_date = datetime.strptime(str(start_t)[:10], "%Y-%m-%d").date()
        e_date = datetime.strptime(str(end_t)[:10], "%Y-%m-%d").date()
        if s_date <= selected_date_obj <= e_date:
          on_leave_employees[cleaned_code] = clean_txt(leave_name)
      except Exception:
        pass

  prev_day = selected_date_obj.strftime("%Y-%m-%d") + " 00:00:00"
  next_day = (selected_date_obj + timedelta(days=1)).strftime(
      "%Y-%m-%d"
  ) + " 05:00:00"

  raw_logs = []
  try:
    logs_res = requests.get(
        f"{BASE_URL}/iclock/api/transactions/?start_time={prev_day}&end_time={next_day}&page_size=5000",
        headers=headers,
        timeout=15,
    )
    if logs_res.status_code == 200:
      raw_logs = logs_res.json().get("data", [])
  except Exception:
    pass

  emp_punches = {}
  for log in raw_logs:
    raw_code = str(log.get("emp_code", "")).strip()
    cleaned_code = str(int(raw_code)) if raw_code.isdigit() else raw_code
    if cleaned_code in active_employees and log.get("punch_time"):
      try:
        p_time = datetime.strptime(
            log.get("punch_time")[:19], "%Y-%m-%d %H:%M:%S"
        )
        dev_sn = str(log.get("terminal_sn", ""))
        dev_name = (
            log.get("terminal_alias")
            or log.get("terminal_name")
            or terminal_map.get(dev_sn, dev_sn or "جهاز رئيسي")
        )
        emp_punches.setdefault(cleaned_code, []).append((p_time, dev_name))
      except Exception:
        continue

  (
      present_staff,
      late_staff,
      absent_staff,
      checkout_staff,
      leave_staff,
      excel_rows,
  ) = ([], [], [], [], [], [])

  for code, emp_data in active_employees.items():
    name = emp_data["name"]
    dept = emp_data["dept"]
    punches = sorted(emp_punches.get(code, []), key=lambda x: x[0])
    filtered_punches = []

    for p_time, d_name in punches:
      if (
          not filtered_punches
          or abs((p_time - filtered_punches[-1][0]).total_seconds()) > 60
      ):
        filtered_punches.append((p_time, d_name))

    day_punches = [
        (p, d)
        for p, d in filtered_punches
        if p.date() == selected_date_obj and p.hour >= 5
    ]

    if not day_punches:
      if code in on_leave_employees:
        leave_reason = on_leave_employees[code]
        leave_staff.append((code, name, dept, leave_reason))
        excel_rows.append({
            "Employee ID": code,
            "First Name": name,
            "Department": dept,
            "Date": selected_date_str,
            "Clock In": "",
            "Clock Out": "",
            "Total WT": "",
            "Status": f"Leave - {leave_reason}",
        })
      else:
        absent_staff.append((code, name, dept))
        excel_rows.append({
            "Employee ID": code,
            "First Name": name,
            "Department": dept,
            "Date": selected_date_str,
            "Clock In": "",
            "Clock Out": "",
            "Total WT": "",
            "Status": "Absence(A)",
        })
      continue

    first_p, first_dev = day_punches[0]

    next_morning = [
        (p, d)
        for p, d in filtered_punches
        if p.date() == selected_date_obj + timedelta(days=1) and p.hour < 5
    ]

    # For exactly one punch, classify an early punch as IN and an afternoon/
    # evening punch as OUT. Calculate the same schedule-aware Actual WT shown
    # by BioTime's Basic Report, without requiring a second upload or report API.
    single_punch_only = len(day_punches) == 1 and not next_morning
    single_punch_is_out = (
        single_punch_only and first_p.hour >= SINGLE_PUNCH_OUT_HOUR
    )

    is_late = (
        not single_punch_is_out
        and (first_p.hour > 9 or (first_p.hour == 9 and first_p.minute > 15))
    )

    if single_punch_only:
      clock_in_value = "" if single_punch_is_out else first_p.strftime("%H:%M")
      clock_out_value = first_p.strftime("%H:%M") if single_punch_is_out else ""
      single_punch_actual_wt = calculate_single_punch_actual_wt(
          first_p,
          single_punch_is_out,
          selected_date_obj,
      )

      if single_punch_is_out:
        checkout_staff.append(
            (code, name, dept, first_p.strftime("%I:%M %p"), first_dev)
        )
        status_str = "Present(P) / Missing IN"
      else:
        present_staff.append(
            (code, name, dept, first_p.strftime("%I:%M %p"), first_dev)
        )
        status_str = (
            "Late(LT) / Missing OUT" if is_late else "Present(P) / Missing OUT"
        )
        if is_late:
          late_staff.append(
              (code, name, dept, first_p.strftime("%I:%M %p"), first_dev)
          )

      excel_rows.append({
          "Employee ID": code,
          "First Name": name,
          "Department": dept,
          "Date": selected_date_str,
          "Clock In": clock_in_value,
          "Clock Out": clock_out_value,
          "Actual WT": single_punch_actual_wt,
          "Calculated WT": single_punch_actual_wt,
          "Total WT": single_punch_actual_wt,
          "Status": status_str,
      })
      continue

    punch_count = (
        2 if (len(day_punches) % 2 != 0 and next_morning) else len(day_punches)
    )

    last_p = None
    last_dev = first_dev

    if punch_count % 2 == 0:
      last_p, last_dev = (
          next_morning[-1]
          if (len(day_punches) % 2 != 0 and next_morning)
          else day_punches[-1]
      )
    elif not is_today and len(day_punches) > 1:
      last_p, last_dev = day_punches[-1]

    ordered_workday_punches = [p for p, _d in day_punches] + [
        p for p, _d in next_morning
    ]
    if len(ordered_workday_punches) > 2 and len(ordered_workday_punches) % 2 == 0:
      total_wt_str = calculate_odd_even_punch_total(ordered_workday_punches)
    else:
      total_wt_str = calculate_total_wt_for_workday(first_p, last_p)

    status_str = "Late(LT)" if is_late else "Present(P)"

    if is_late:
      late_staff.append(
          (code, name, dept, first_p.strftime("%I:%M %p"), first_dev)
      )

    if is_today:
      if punch_count % 2 != 0:
        present_staff.append(
            (code, name, dept, first_p.strftime("%I:%M %p"), first_dev)
        )
        excel_rows.append({
            "Employee ID": code,
            "First Name": name,
            "Department": dept,
            "Date": selected_date_str,
            "Clock In": first_p.strftime("%H:%M"),
            "Clock Out": "",
            "Total WT": "",
            "Status": status_str,
        })
      else:
        last_p_real, last_dev_real = (
            next_morning[-1]
            if (len(day_punches) % 2 != 0 and next_morning)
            else day_punches[-1]
        )
        checkout_staff.append((
            code,
            name,
            dept,
            last_p_real.strftime("%I:%M %p"),
            last_dev_real,
        ))
        excel_rows.append({
            "Employee ID": code,
            "First Name": name,
            "Department": dept,
            "Date": selected_date_str,
            "Clock In": first_p.strftime("%H:%M"),
            "Clock Out": last_p_real.strftime("%H:%M"),
            "Total WT": total_wt_str,
            "Status": status_str,
        })
    else:
      if last_p:
        checkout_staff.append(
            (code, name, dept, last_p.strftime("%I:%M %p"), last_dev)
        )
      else:
        present_staff.append(
            (code, name, dept, first_p.strftime("%I:%M %p"), first_dev)
        )

      excel_rows.append({
          "Employee ID": code,
          "First Name": name,
          "Department": dept,
          "Date": selected_date_str,
          "Clock In": first_p.strftime("%H:%M"),
          "Clock Out": last_p.strftime("%H:%M") if last_p else "",
          "Total WT": total_wt_str,
          "Status": status_str,
      })

  absent_staff.sort(key=lambda x: int(x[0]) if x[0].isdigit() else 999)
  present_staff.sort(key=lambda x: int(x[0]) if x[0].isdigit() else 999)
  late_staff.sort(key=lambda x: int(x[0]) if x[0].isdigit() else 999)
  leave_staff.sort(key=lambda x: int(x[0]) if x[0].isdigit() else 999)
  checkout_staff.sort(key=lambda x: int(x[0]) if x[0].isdigit() else 999)

  return (
      active_employees,
      present_staff,
      late_staff,
      absent_staff,
      checkout_staff,
      leave_staff,
      devices,
      excel_rows,
  )



# ==========================================
# 2B. DIRECT MONTHLY REPORT WORKBOOK
# ==========================================
def _monthly_minutes(value):
  """Convert HH:MM work time to minutes for report summaries."""
  if value is None:
    return 0
  raw = str(value).strip()
  if not raw or ":" not in raw:
    return 0
  try:
    hours, minutes = raw.split(":", 1)[:2]
    return max(0, int(hours) * 60 + int(minutes[:2]))
  except (TypeError, ValueError):
    return 0


def _monthly_hhmm(total_minutes):
  total_minutes = max(0, int(total_minutes or 0))
  return f"{total_minutes // 60}:{total_minutes % 60:02d}"


def _monthly_status_fields(attendance):
  status = str(attendance.get("Status", "") or "")
  clock_in = str(attendance.get("Clock In", "") or "").strip()
  clock_out = str(attendance.get("Clock Out", "") or "").strip()
  if "Leave" in status:
    category = "Leave"
  elif "Absence" in status:
    category = "Absent"
  elif "Late" in status:
    category = "Late"
  else:
    category = "Present"
  single = "Yes" if bool(clock_in) != bool(clock_out) and category not in ("Leave", "Absent") else "No"
  multi = "Yes" if bool(attendance.get("Odd Even Paired")) else "No"
  if single == "Yes":
    rule = "Single Punch → Actual WT"
  elif multi == "Yes":
    rule = "Odd/Even punch pairs"
  elif category not in ("Leave", "Absent"):
    rule = "Normal IN/OUT → Total WT"
  else:
    rule = category
  return category, single, multi, rule


def _monthly_apply_header(ws, row=1):
  header_fill = PatternFill("solid", fgColor="0F172A")
  for cell in ws[row]:
    cell.fill = header_fill
    cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
  ws.row_dimensions[row].height = 26
  ws.freeze_panes = f"A{row + 1}"
  ws.sheet_view.showGridLines = False


def _monthly_add_table(ws, name):
  if ws.max_row < 2 or ws.max_column < 1:
    return
  ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
  table = Table(displayName=name, ref=ref)
  table.tableStyleInfo = TableStyleInfo(
      name="TableStyleMedium2",
      showFirstColumn=False,
      showLastColumn=False,
      showRowStripes=True,
      showColumnStripes=False,
  )
  ws.add_table(table)


def _monthly_autowidth(ws, max_width=34):
  for column_cells in ws.columns:
    letter = get_column_letter(column_cells[0].column)
    width = 10
    for cell in column_cells[:250]:
      if cell.value is None:
        continue
      width = max(width, min(max_width, len(str(cell.value)) + 2))
    ws.column_dimensions[letter].width = width


def build_monthly_attendance_workbook(all_attendance, employee_details, start_date, end_date, raw_transactions=None):
  """Create a tidy direct-from-BioTime monthly workbook plus UI summary data."""
  raw_transactions = raw_transactions or []
  dates = []
  cursor = start_date
  while cursor <= end_date:
    dates.append(cursor)
    cursor += timedelta(days=1)

  employee_ids = sorted(
      employee_details.keys(),
      key=lambda value: (0, int(value)) if str(value).isdigit() else (1, str(value)),
  )

  daily_rows = []
  exception_rows = []
  total_minutes_ui = 0
  working_days_ui = 0
  late_ui = 0
  absent_ui = 0
  leave_ui = 0
  single_ui = 0

  for attendance_date in dates:
    date_map = all_attendance.get(attendance_date, {})
    for employee_id in employee_ids:
      info = employee_details.get(employee_id, {})
      attendance = date_map.get(employee_id) or {
          "Employee ID": employee_id,
          "First Name": info.get("name", ""),
          "Department": info.get("dept", "غير محدد"),
          "Date": attendance_date.strftime("%Y-%m-%d"),
          "Clock In": "",
          "Clock Out": "",
          "Calculated WT": "",
          "Status": "Absence(A)",
          "Punch Count": 0,
          "Odd Even Paired": False,
      }
      category, single, multi, rule = _monthly_status_fields(attendance)
      work_text = str(
          attendance.get("Calculated WT", "")
          or attendance.get("Actual WT", "")
          or attendance.get("Total WT", "")
          or ""
      ).strip()
      work_minutes = _monthly_minutes(work_text)
      if category in ("Present", "Late"):
        working_days_ui += 1
        total_minutes_ui += work_minutes
      if category == "Late":
        late_ui += 1
      elif category == "Absent":
        absent_ui += 1
      elif category == "Leave":
        leave_ui += 1
      if single == "Yes":
        single_ui += 1

      row = {
          "BioTime ID": employee_id,
          "Employee": info.get("name", attendance.get("First Name", "")),
          "Department": info.get("dept", attendance.get("Department", "غير محدد")),
          "Date": attendance_date,
          "Clock In": attendance.get("Clock In", ""),
          "Clock Out": attendance.get("Clock Out", ""),
          "Working Hours": work_minutes / 1440 if work_minutes else None,
          "Working Hours Text": work_text,
          "Day Status": category,
          "Single Punch": single,
          "Multi Punch": multi,
          "Punch Count": int(attendance.get("Punch Count", 0) or 0),
          "Calculation Rule": rule,
          "BioTime Status": str(attendance.get("Status", "") or ""),
      }
      daily_rows.append(row)
      if category in ("Late", "Absent", "Leave") or single == "Yes" or multi == "Yes":
        exception_rows.append(row.copy())

  wb = openpyxl.Workbook()
  ws_overview = wb.active
  ws_overview.title = "Overview"
  ws_overview.sheet_view.showGridLines = False
  ws_overview.merge_cells("A1:F1")
  ws_overview["A1"] = "BioTime Monthly Attendance"
  ws_overview["A1"].font = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
  ws_overview["A1"].fill = PatternFill("solid", fgColor="1D4ED8")
  ws_overview["A1"].alignment = Alignment(horizontal="center", vertical="center")
  ws_overview.row_dimensions[1].height = 34
  ws_overview.merge_cells("A2:F2")
  ws_overview["A2"] = f"{start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
  ws_overview["A2"].font = Font(name="Calibri", size=10, color="475569")
  ws_overview["A2"].alignment = Alignment(horizontal="center")

  overview_labels = [
      ("A4", "Employees", "B4", "=COUNTA('Monthly Summary'!A2:A1000)"),
      ("C4", "Working Days", "D4", "=SUM('Monthly Summary'!D2:D1000)"),
      ("E4", "Total Hours", "F4", "=SUM('Monthly Summary'!I2:I1000)"),
      ("A6", "Late", "B6", "=SUM('Monthly Summary'!E2:E1000)"),
      ("C6", "Absent", "D6", "=SUM('Monthly Summary'!F2:F1000)"),
      ("E6", "Single Punch", "F6", "=SUM('Monthly Summary'!H2:H1000)"),
  ]
  for label_cell, label, value_cell, formula in overview_labels:
    ws_overview[label_cell] = label
    ws_overview[label_cell].font = Font(bold=True, color="64748B")
    ws_overview[value_cell] = formula
    ws_overview[value_cell].font = Font(bold=True, size=16, color="0F172A")
    for coord in (label_cell, value_cell):
      ws_overview[coord].fill = PatternFill("solid", fgColor="F8FAFC")
      ws_overview[coord].alignment = Alignment(horizontal="center", vertical="center")
  ws_overview["F4"].number_format = "[h]:mm"
  for col in range(1, 7):
    ws_overview.column_dimensions[get_column_letter(col)].width = 18
  ws_overview["A9"] = "Calculation rules"
  ws_overview["A9"].font = Font(bold=True, color="0F172A")
  ws_overview["A10"] = "Single punch"
  ws_overview["B10"] = "Actual WT"
  ws_overview["A11"] = "Normal IN/OUT"
  ws_overview["B11"] = "Total WT"
  ws_overview["A12"] = "4/6/8... punches"
  ws_overview["B12"] = "Sequential odd/even pairs"
  ws_overview["A14"] = "Generated"
  ws_overview["B14"] = datetime.now(SYRIA_TZ).strftime("%d/%m/%Y %H:%M")

  # Detailed normalized data.
  ws_daily = wb.create_sheet("Daily Attendance")
  daily_headers = list(daily_rows[0].keys()) if daily_rows else [
      "BioTime ID", "Employee", "Department", "Date", "Clock In", "Clock Out",
      "Working Hours", "Working Hours Text", "Day Status", "Single Punch",
      "Multi Punch", "Punch Count", "Calculation Rule", "BioTime Status"
  ]
  ws_daily.append(daily_headers)
  for row in daily_rows:
    ws_daily.append([row.get(header) for header in daily_headers])
  _monthly_apply_header(ws_daily)
  for row_no in range(2, ws_daily.max_row + 1):
    ws_daily.cell(row_no, 4).number_format = "dd/mm/yyyy"
    ws_daily.cell(row_no, 7).number_format = "[h]:mm"
    for col_no in (1, 2, 3, 4, 5, 6, 8, 9, 10, 11, 12, 13, 14):
      ws_daily.cell(row_no, col_no).font = Font(name="Calibri", size=10, color="008000")
  _monthly_add_table(ws_daily, "DailyAttendanceTable")
  _monthly_autowidth(ws_daily)

  # Formula-driven employee summary.
  ws_summary = wb.create_sheet("Monthly Summary")
  summary_headers = [
      "BioTime ID", "Employee", "Department", "Working Days", "Late Days",
      "Absent Days", "Leave Days", "Single Punch", "Total Hours"
  ]
  ws_summary.append(summary_headers)
  for idx, employee_id in enumerate(employee_ids, start=2):
    info = employee_details.get(employee_id, {})
    ws_summary.cell(idx, 1, employee_id)
    ws_summary.cell(idx, 2, info.get("name", ""))
    ws_summary.cell(idx, 3, info.get("dept", "غير محدد"))
    # Daily Attendance: A=ID, G=Hours, I=Day Status, J=Single Punch.
    ws_summary.cell(idx, 4, f'=COUNTIFS(\'Daily Attendance\'!$A:$A,A{idx},\'Daily Attendance\'!$I:$I,"Present")+COUNTIFS(\'Daily Attendance\'!$A:$A,A{idx},\'Daily Attendance\'!$I:$I,"Late")')
    ws_summary.cell(idx, 5, f'=COUNTIFS(\'Daily Attendance\'!$A:$A,A{idx},\'Daily Attendance\'!$I:$I,"Late")')
    ws_summary.cell(idx, 6, f'=COUNTIFS(\'Daily Attendance\'!$A:$A,A{idx},\'Daily Attendance\'!$I:$I,"Absent")')
    ws_summary.cell(idx, 7, f'=COUNTIFS(\'Daily Attendance\'!$A:$A,A{idx},\'Daily Attendance\'!$I:$I,"Leave")')
    ws_summary.cell(idx, 8, f'=COUNTIFS(\'Daily Attendance\'!$A:$A,A{idx},\'Daily Attendance\'!$J:$J,"Yes")')
    ws_summary.cell(idx, 9, f'=SUMIFS(\'Daily Attendance\'!$G:$G,\'Daily Attendance\'!$A:$A,A{idx})')
    ws_summary.cell(idx, 9).number_format = "[h]:mm"
    for col_no in range(4, 10):
      ws_summary.cell(idx, col_no).font = Font(name="Calibri", size=10, color="000000")
    for col_no in range(1, 4):
      ws_summary.cell(idx, col_no).font = Font(name="Calibri", size=10, color="008000")
  _monthly_apply_header(ws_summary)
  _monthly_add_table(ws_summary, "MonthlySummaryTable")
  _monthly_autowidth(ws_summary)

  # Calendar/pivot-style matrix.
  ws_matrix = wb.create_sheet("Calendar Matrix")
  matrix_headers = ["BioTime ID", "Employee", "Department"] + [d.strftime("%d/%m") for d in dates] + ["Days", "Total Hours"]
  ws_matrix.append(matrix_headers)
  for row_no, employee_id in enumerate(employee_ids, start=2):
    info = employee_details.get(employee_id, {})
    ws_matrix.cell(row_no, 1, employee_id)
    ws_matrix.cell(row_no, 2, info.get("name", ""))
    ws_matrix.cell(row_no, 3, info.get("dept", "غير محدد"))
    for date_index, attendance_date in enumerate(dates, start=4):
      attendance = all_attendance.get(attendance_date, {}).get(employee_id, {})
      category, _single, _multi, _rule = _monthly_status_fields(attendance)
      work_text = str(attendance.get("Calculated WT", "") or attendance.get("Actual WT", "") or attendance.get("Total WT", "") or "").strip()
      work_minutes = _monthly_minutes(work_text)
      cell = ws_matrix.cell(row_no, date_index)
      if category == "Absent":
        cell.value = "A"
        cell.fill = PatternFill("solid", fgColor="FEE2E2")
        cell.font = Font(color="991B1B", bold=True)
      elif category == "Leave":
        cell.value = "L"
        cell.fill = PatternFill("solid", fgColor="E0F2FE")
        cell.font = Font(color="0369A1", bold=True)
      elif work_minutes:
        cell.value = work_minutes / 1440
        cell.number_format = "[h]:mm"
        if category == "Late":
          cell.fill = PatternFill("solid", fgColor="FEF3C7")
      else:
        cell.value = None
      cell.alignment = Alignment(horizontal="center", vertical="center")
    first_day_col = 4
    last_day_col = 3 + len(dates)
    days_col = last_day_col + 1
    total_col = last_day_col + 2
    first_letter = get_column_letter(first_day_col)
    last_letter = get_column_letter(last_day_col)
    ws_matrix.cell(row_no, days_col, f'=COUNT({first_letter}{row_no}:{last_letter}{row_no})')
    ws_matrix.cell(row_no, total_col, f'=SUM({first_letter}{row_no}:{last_letter}{row_no})')
    ws_matrix.cell(row_no, total_col).number_format = "[h]:mm"
  _monthly_apply_header(ws_matrix)
  ws_matrix.freeze_panes = "D2"
  ws_matrix.column_dimensions["A"].width = 12
  ws_matrix.column_dimensions["B"].width = 26
  ws_matrix.column_dimensions["C"].width = 22
  for col_no in range(4, 4 + len(dates)):
    ws_matrix.column_dimensions[get_column_letter(col_no)].width = 9
  if dates:
    ws_matrix.column_dimensions[get_column_letter(4 + len(dates))].width = 10
    ws_matrix.column_dimensions[get_column_letter(5 + len(dates))].width = 13

  # Exceptions only.
  ws_exc = wb.create_sheet("Exceptions")
  exc_headers = daily_headers
  ws_exc.append(exc_headers)
  for row in exception_rows:
    ws_exc.append([row.get(header) for header in exc_headers])
  _monthly_apply_header(ws_exc)
  for row_no in range(2, ws_exc.max_row + 1):
    ws_exc.cell(row_no, 4).number_format = "dd/mm/yyyy"
    ws_exc.cell(row_no, 7).number_format = "[h]:mm"
  if exception_rows:
    _monthly_add_table(ws_exc, "AttendanceExceptionsTable")
  _monthly_autowidth(ws_exc)

  # Raw transactions for audit/recovery.
  ws_raw = wb.create_sheet("Raw Transactions")
  raw_headers = ["BioTime ID", "Punch Time", "Punch State", "Device", "Terminal SN", "Source ID"]
  ws_raw.append(raw_headers)
  for item in raw_transactions:
    employee_id = item.get("emp_code")
    employee_obj = item.get("employee") if isinstance(item.get("employee"), dict) else {}
    employee_id = employee_id or employee_obj.get("emp_code") or item.get("employee_code") or ""
    ws_raw.append([
        str(employee_id or ""),
        item.get("punch_time") or item.get("punch_datetime") or item.get("timestamp") or "",
        item.get("punch_state") if item.get("punch_state") is not None else item.get("state", ""),
        item.get("terminal_alias") or item.get("terminal_name") or "",
        item.get("terminal_sn") or item.get("sn") or "",
        item.get("id") or "",
    ])
  _monthly_apply_header(ws_raw)
  if raw_transactions:
    _monthly_add_table(ws_raw, "RawTransactionsTable")
  _monthly_autowidth(ws_raw)

  # Workbook calculation flags so formulas refresh in Excel.
  try:
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
  except Exception:
    pass

  output = io.BytesIO()
  wb.save(output)
  output.seek(0)

  ui_summary = {
      "employees": len(employee_ids),
      "working_days": working_days_ui,
      "total_minutes": total_minutes_ui,
      "total_hours": _monthly_hhmm(total_minutes_ui),
      "late": late_ui,
      "absent": absent_ui,
      "leave": leave_ui,
      "single_punch": single_ui,
      "from": start_date.strftime("%d/%m/%Y"),
      "to": end_date.strftime("%d/%m/%Y"),
  }
  return output.getvalue(), ui_summary, exception_rows


# ==========================================
# 3. INTERFACE RENDERING
# ==========================================
now_syria = datetime.now(SYRIA_TZ)
today_str = now_syria.strftime("%Y-%m-%d")

strlit.markdown(
    '<div class="gp-app-hero">'
    '<div class="gp-app-hero-title">📡 Golden Palace • BioTime</div>'
    '<div class="gp-app-hero-subtitle">حضور يومي • تقرير شهري • نسخة احتياطية</div>'
    '</div>',
    unsafe_allow_html=True,
)

strlit.markdown('<div class="gp-top-controls"></div>', unsafe_allow_html=True)
c_date, c_status, c_ref = strlit.columns([1.35, 0.90, 1.00], gap="small")
with c_date:
  strlit.markdown('<div class="gp-control-caption">📅 تاريخ العرض</div>', unsafe_allow_html=True)
  selected_date_obj_input = strlit.date_input(
      "", value=now_syria.date(), label_visibility="collapsed"
  )
  selected_date_str = selected_date_obj_input.strftime("%Y-%m-%d")

is_today = selected_date_str == today_str

with c_status:
  strlit.markdown('<div class="gp-control-caption">📡 حالة الاتصال</div>', unsafe_allow_html=True)
  if is_today:
    strlit.markdown(
        '<div class="gp-live-card">'
        '<span class="gp-live-dot"></span>'
        '<span class="gp-live-main">Online</span>'
        '<span class="gp-live-sub">مباشر</span>'
        '</div>',
        unsafe_allow_html=True,
    )
  else:
    strlit.markdown(
        f'<div class="gp-archive-card">'
        f'<span>🗂️</span><span class="gp-live-main">أرشيف</span>'
        f'<span class="gp-live-sub">{selected_date_obj_input.strftime("%d/%m/%Y")}</span>'
        f'</div>',
        unsafe_allow_html=True,
    )

with c_ref:
  strlit.markdown('<div class="gp-control-caption">⚡ البيانات</div>', unsafe_allow_html=True)
  if strlit.button("🔄 تحديث الآن", use_container_width=True, key="top_refresh_button"):
    strlit.cache_data.clear()
    strlit.rerun()

if "last_selected_date" not in strlit.session_state:
  strlit.session_state["last_selected_date"] = selected_date_str

if strlit.session_state["last_selected_date"] != selected_date_str:
  strlit.session_state["last_selected_date"] = selected_date_str
  strlit.session_state["selected_view"] = "present" if is_today else "all"

main_loading_overlay = show_loading_overlay("جاري تحديث بيانات BioTime...")
try:
  act, pre, lat, abs_s, chk, lev, devices, exc = load_attendance_data_from_api(
      selected_date_str, selected_date_obj_input, is_today
  )
  hide_loading_overlay(main_loading_overlay)
  main_loading_overlay = None

  daily_tab, monthly_tab, backup_tab = strlit.tabs([
      "📅 الحضور اليومي",
      "📊 التقرير الشهري",
      "🛡️ النسخة الاحتياطية",
  ])

  with daily_tab:
    render_clickable_attendance_cards(act, pre, lat, chk, lev, abs_s)
    # 📥 UPLOAD TEMPLATE & FILL ATTENDANCE VALUES OR GENERATE DEFAULT REPORT
    col_gen, col_up = strlit.columns(2)

    with col_gen:
      # Standard Generated Attendance Report
      df_excel = pd.DataFrame(exc)
      output = io.BytesIO()

      wb = openpyxl.Workbook()
      ws = wb.active
      ws.title = "Attendance Report"

      thin_border = Border(
          left=Side(style="thin", color="D3D3D3"),
          right=Side(style="thin", color="D3D3D3"),
          top=Side(style="thin", color="D3D3D3"),
          bottom=Side(style="thin", color="D3D3D3"),
      )

      headers = [
          "Employee ID",
          "First Name",
          "Department",
          "Date",
          "Clock In",
          "Clock Out",
          "Total WT",
          "Status",
      ]
      ws.append(headers)
      ws.row_dimensions[1].height = 24

      for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="1F4E78", end_color="1F4E78", fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

      for idx, row_data in enumerate(exc, 2):
        ws.row_dimensions[idx].height = 20
        ws.append([
            row_data["Employee ID"],
            row_data["First Name"],
            row_data["Department"],
            row_data["Date"],
            row_data["Clock In"],
            row_data["Clock Out"],
            row_data["Total WT"],
            row_data["Status"],
        ])

        status_val = str(row_data["Status"])
        row_fill = None
        row_font_color = "000000"

        if "Leave" in status_val or "L" in status_val:
          row_fill = PatternFill(
              start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
          )
          row_font_color = "002060"
        elif "Absence" in status_val or "A" in status_val:
          row_fill = PatternFill(
              start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
          )
          row_font_color = "9C0006"

        for col_idx in range(1, 9):
          cell = ws.cell(row=idx, column=col_idx)

          if row_fill:
            cell.fill = row_fill
            cell.font = Font(
                name="Calibri", size=11, bold=True, color=row_font_color
            )
          else:
            cell.font = Font(name="Calibri", size=11)

          if col_idx == 8 and not row_fill:
            if "Late" in status_val or "LT" in status_val:
              cell.font = Font(
                  name="Calibri", size=11, bold=True, color="9C0006"
              )
              cell.fill = PatternFill(
                  start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
              )
            elif "Present" in status_val or "P" in status_val:
              cell.font = Font(
                  name="Calibri", size=11, bold=True, color="006100"
              )
              cell.fill = PatternFill(
                  start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
              )

          cell.alignment = Alignment(horizontal="center", vertical="center")
          cell.border = thin_border

      for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
          if cell.value:
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 5, 14)

      wb.save(output)
      excel_data = output.getvalue()

      strlit.download_button(
          label="📥 تحميل تقرير اليوم (افتراضي)",
          data=excel_data,
          file_name=f"Daily_Attendance_Report_{selected_date_str}.xlsx",
          mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
          use_container_width=True,
      )

    with col_up:
      strlit.markdown(
          '<div class="gp-section-title">📊 جدول الدوام الشهري</div>'
          '<div class="gp-file-note">ارفع ملف Excel فقط — بيانات BioTime تُجلب تلقائياً.</div>',
          unsafe_allow_html=True,
      )
      # ONE upload box only: the monthly attendance Excel template.
      # BioTime calculated work hours are fetched automatically from BioTime.
      uploaded_template = strlit.file_uploader(
          "📂 رفع جدول الدوام الشهري وتعبئته تلقائياً",
          type=["xlsx", "xlsm"],
          label_visibility="collapsed",
          key="monthly_attendance_template",
      )
      strlit.markdown(
          '<div class="gp-tech-note">Single Punch → Actual WT &nbsp;|&nbsp; Normal IN/OUT → Total WT</div>',
          unsafe_allow_html=True,
      )

      def normalize_id(value):
        """Normalize an employee ID without changing its identity."""
        if value is None:
          return ""

        # Excel may give a numeric ID as 1.0 when the source cell is numeric.
        if isinstance(value, float) and value.is_integer():
          value = int(value)

        raw = str(value).strip()
        if not raw:
          return ""

        if raw.endswith(".0") and raw[:-2].isdigit():
          raw = raw[:-2]

        return raw

      def normalize_report_key(value):
        """Normalize report JSON field names for tolerant matching."""
        return re.sub(r"[^a-z0-9]", "", str(value or "").lower())

      def flatten_report_row(value, prefix=""):
        """Flatten nested report JSON while retaining short and full keys."""
        flat = {}

        if isinstance(value, dict):
          for key, child in value.items():
            normalized_key = normalize_report_key(key)
            full_key = f"{prefix}{normalized_key}" if prefix else normalized_key

            if isinstance(child, (dict, list)):
              flat.update(flatten_report_row(child, full_key))
            else:
              if normalized_key and normalized_key not in flat:
                flat[normalized_key] = child
              if full_key:
                flat[full_key] = child

        elif isinstance(value, list):
          for item in value:
            flat.update(flatten_report_row(item, prefix))

        return flat

      def extract_report_rows(payload):
        """Extract attendance rows from BioTime report JSON.

        BioTime releases return report data in several shapes. Some are flat DRF
        lists, while others group daily rows beneath an employee object. Preserve
        scalar parent fields while walking nested lists so employee/date context is
        not lost before normalization.
        """
        rows = []
        seen = set()

        row_hint_keys = {
            "attdate",
            "attendancedate",
            "workdate",
            "date",
            "clockin",
            "clockout",
            "checkin",
            "checkout",
            "actualwt",
            "actualworked",
            "actualworktime",
            "totalwt",
            "totalworked",
            "totalworktime",
        }

        def walk(value, inherited=None):
          inherited = dict(inherited or {})

          if isinstance(value, list):
            for item in value:
              walk(item, inherited)
            return

          if not isinstance(value, dict):
            return

          scalar_context = dict(inherited)
          nested_items = []

          for key, child in value.items():
            if isinstance(child, (dict, list)):
              nested_items.append((key, child))
              # Employee/report group metadata is often a nested object sitting
              # beside the actual daily rows. Carry its scalar fields into sibling
              # row context using prefixed names such as employee_emp_code.
              if isinstance(child, dict):
                for child_key, child_value in child.items():
                  if not isinstance(child_value, (dict, list)):
                    scalar_context.setdefault(
                        f"{key}_{child_key}",
                        child_value,
                    )
            else:
              scalar_context.setdefault(key, child)

          normalized_keys = {normalize_report_key(key) for key in value.keys()}
          has_row_hint = bool(normalized_keys & row_hint_keys)

          if has_row_hint:
            merged = dict(inherited)
            merged.update(value)
            # Avoid adding the same object twice when several envelope paths point
            # to the identical dictionary.
            signature = id(value)
            if signature not in seen:
              seen.add(signature)
              rows.append(merged)

          for _key, child in nested_items:
            walk(child, scalar_context)

        walk(payload)

        # Flat report lists occasionally contain fields whose names are unknown to
        # row_hint_keys. Fall back to the conventional envelopes if needed.
        if rows:
          return rows

        if isinstance(payload, list):
          return [item for item in payload if isinstance(item, dict)]
        if isinstance(payload, dict):
          for key in ("data", "results", "rows", "items"):
            candidate = payload.get(key)
            if isinstance(candidate, list):
              return [item for item in candidate if isinstance(item, dict)]
        return []

      def first_report_value(flat_row, candidate_keys):
        """Return the first non-empty report field, preferring exact keys."""
        normalized_candidates = [
            normalize_report_key(candidate_key) for candidate_key in candidate_keys
        ]

        # Exact keys first. flatten_report_row keeps short child keys, so this
        # avoids mistaking e.g. total_work_time for the more specific work_time.
        for normalized_candidate in normalized_candidates:
          if not normalized_candidate:
            continue
          value = flat_row.get(normalized_candidate)
          if value not in (None, ""):
            return value, normalized_candidate

        # Fallback for unusual nested serializers that expose only prefixed keys.
        for normalized_candidate in normalized_candidates:
          if not normalized_candidate:
            continue
          for key, value in flat_row.items():
            if value in (None, ""):
              continue
            if key.endswith(normalized_candidate):
              return value, key
        return "", ""

      def parse_report_date(value, fallback_date=None):
        if isinstance(value, datetime):
          return value.date()
        if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
          return value
        if value not in (None, ""):
          raw = str(value).strip()

          # BioTime versions may serialize dates as YYYY-MM-DD, full timestamps,
          # or ISO-8601 strings with a T/Z timezone marker.
          try:
            return datetime.fromisoformat(raw.replace("Z", "+00:00")).date()
          except ValueError:
            pass

          for candidate, fmt in (
              (raw[:10], "%Y-%m-%d"),
              (raw[:19], "%Y-%m-%d %H:%M:%S"),
              (raw[:10], "%d/%m/%Y"),
              (raw[:10], "%d-%m-%Y"),
              (raw[:8], "%d/%m/%y"),
              (raw[:8], "%d-%m-%y"),
          ):
            try:
              return datetime.strptime(candidate, fmt).date()
            except ValueError:
              continue
        return fallback_date

      def duration_to_hhmm(value, source_key=""):
        """Normalize a BioTime calculated duration to HH:MM."""
        if value is None or str(value).strip() == "":
          return ""

        if isinstance(value, timedelta):
          total_seconds = int(value.total_seconds())
        elif isinstance(value, (int, float)) and not isinstance(value, bool):
          number = float(value)
          key = normalize_report_key(source_key)
          if "second" in key:
            total_seconds = round(number)
          elif "minute" in key:
            total_seconds = round(number * 60)
          elif 0 <= number <= 1 and not number.is_integer():
            # Some report serializers return an Excel/day fraction.
            total_seconds = round(number * 86400)
          elif 0 <= number <= 24:
            total_seconds = round(number * 3600)
          elif 0 <= number <= 1440:
            total_seconds = round(number * 60)
          else:
            total_seconds = round(number)
        else:
          raw = str(value).strip()

          day_match = re.match(
              r"^(?P<days>\d+)\s+day[s]?,\s*(?P<hours>\d+):(?P<minutes>\d{1,2})(?::(?P<seconds>\d{1,2}))?$",
              raw,
              re.IGNORECASE,
          )
          if day_match:
            total_seconds = (
                int(day_match.group("days")) * 86400
                + int(day_match.group("hours")) * 3600
                + int(day_match.group("minutes")) * 60
                + int(day_match.group("seconds") or 0)
            )
          else:
            time_match = re.match(
                r"^(?P<hours>\d+):(?P<minutes>\d{1,2})(?::(?P<seconds>\d{1,2}))?$",
                raw,
            )
            if time_match:
              total_seconds = (
                  int(time_match.group("hours")) * 3600
                  + int(time_match.group("minutes")) * 60
                  + int(time_match.group("seconds") or 0)
              )
            else:
              try:
                number = float(raw.replace(",", "."))
              except ValueError:
                return ""
              total_seconds = round(number * 3600)

        if total_seconds < 0:
          return ""

        hours, remainder = divmod(total_seconds, 3600)
        minutes = remainder // 60
        return f"{hours:02d}:{minutes:02d}"

      def normalize_clock_value(value):
        """Normalize a report clock value to HH:MM without inventing a punch."""
        if value in (None, ""):
          return ""
        if isinstance(value, datetime):
          return value.strftime("%H:%M")
        if hasattr(value, "hour") and hasattr(value, "minute"):
          try:
            return f"{int(value.hour):02d}:{int(value.minute):02d}"
          except Exception:
            pass

        raw = str(value).strip()
        if not raw:
          return ""

        for fmt in (
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%d-%m-%Y %H:%M:%S",
            "%d/%m/%Y %H:%M:%S",
            "%H:%M:%S",
            "%H:%M",
            "%I:%M %p",
        ):
          try:
            return datetime.strptime(raw[:19], fmt).strftime("%H:%M")
          except ValueError:
            continue

        # Keep an already-recognisable HH:MM prefix.
        match = re.search(r"(?<!\d)(\d{1,2}):(\d{2})(?::\d{2})?", raw)
        if match:
          try:
            hour = int(match.group(1))
            minute = int(match.group(2))
            if 0 <= hour <= 23 and 0 <= minute <= 59:
              return f"{hour:02d}:{minute:02d}"
          except ValueError:
            pass
        return raw

      def normalize_calculated_report_row(
          report_row,
          start_date,
          end_date,
          employee_internal_id_map=None,
      ):
        """Convert one BioTime attendance-report row to the app attendance shape.

        The BioTime report APIs are not fully consistent between releases. Some
        versions return emp_code directly; others return an internal employee id or
        a nested employee object. The same is true for worked-time field names.
        """
        flat = flatten_report_row(report_row)
        employee_internal_id_map = employee_internal_id_map or {}

        employee_value, _employee_key = first_report_value(
            flat,
            (
                "emp_code",
                "employee_emp_code",
                "employee_code",
                "employee_code_code",
                "employeecode",
                "empcode",
                "staff_code",
                "staffcode",
                "employee_number",
                "employee_no",
            ),
        )
        employee_id = normalize_id(employee_value)

        if not employee_id:
          internal_value, _internal_key = first_report_value(
              flat,
              (
                  "employee_id",
                  "employeeid",
                  "emp_id",
                  "empid",
                  "employee",
                  "emp",
              ),
          )
          internal_id = normalize_id(internal_value)
          employee_id = employee_internal_id_map.get(internal_id, "")

        if not employee_id:
          return None

        date_value, _date_key = first_report_value(
            flat,
            (
                "att_date",
                "attendance_date",
                "attendance_day",
                "work_date",
                "workday",
                "attdate",
                "transaction_date",
                "date",
            ),
        )
        attendance_date = parse_report_date(date_value)
        if attendance_date is None or not (start_date <= attendance_date <= end_date):
          return None

        clock_in, _clock_in_key = first_report_value(
            flat,
            (
                "clock_in",
                "check_in",
                "checkin",
                "first_in",
                "first_punch",
                "first_punch_time",
                "checkin_time",
                "in_time",
                "punch_in",
                "first_clock_in",
            ),
        )
        clock_out, _clock_out_key = first_report_value(
            flat,
            (
                "clock_out",
                "check_out",
                "checkout",
                "last_out",
                "last_punch",
                "last_punch_time",
                "checkout_time",
                "out_time",
                "punch_out",
                "last_clock_out",
            ),
        )

        # Keep Actual WT and Total WT separate. Generic "working/worked hours"
        # fields belong to Total WT (the uncapped interval shown by BioTime's
        # Monthly Attendance Summary), not Actual WT. This distinction is critical:
        # single-punch days use Actual WT; normal IN/OUT days use Total WT.
        actual_value, actual_key = first_report_value(
            flat,
            (
                "actual_wt",
                "actualwt",
                "actual_work_time",
                "actual_working_time",
                "actual_worked_time",
                "actual_worked_hours",
                "actual_worked",
                "actual_work",
                "actual_time",
                "actual_duration",
                "actual_work_duration",
                "actual_worked_minutes",
                "actual_worked_seconds",
            ),
        )
        actual_work_time = duration_to_hhmm(actual_value, actual_key)

        total_value, total_key = first_report_value(
            flat,
            (
                "total_wt",
                "totalwt",
                "total_worked",
                "total_work_time",
                "total_worked_time",
                "total_worked_hours",
                "total_duration",
                "total_hours",
                "total_hrs",
                "total_time",
                "working_hrs",
                "working_hours",
                "working_time",
                "work_hrs",
                "work_hours",
                "work_time",
                "worked_hrs",
                "worked_hours",
                "worked_time",
                "worked_duration",
                "work_duration",
                "worked",
                "duration",
            ),
        )
        total_work_time = duration_to_hhmm(total_value, total_key)
        calculated_work_time = total_work_time or actual_work_time

        clock_in = normalize_clock_value(clock_in)
        clock_out = normalize_clock_value(clock_out)

        # Ignore report metadata rows that do not describe attendance.
        if not calculated_work_time and not clock_in and not clock_out:
          return None

        if clock_in and not clock_out:
          status = "Present(P) / Missing OUT"
        elif clock_out and not clock_in:
          status = "Present(P) / Missing IN"
        elif calculated_work_time:
          status = "Present(P)"
        else:
          status = ""

        return {
            "Employee ID": employee_id,
            "Attendance Date": attendance_date,
            "Date": attendance_date.strftime("%Y-%m-%d"),
            "Clock In": clock_in,
            "Clock Out": clock_out,
            "Actual WT": actual_work_time,
            "Report Total WT": total_work_time,
            "Calculated WT": calculated_work_time,
            "Total WT": total_work_time,
            "Status": status,
        }

      def report_record_score(record, expected_single_punch=False):
        """Prefer the report row most useful for single-punch recovery."""
        score = 0
        if expected_single_punch:
          score += 100
        if record.get("Actual WT"):
          score += 50
        if record.get("Report Total WT"):
          score += 30
        if record.get("Calculated WT"):
          score += 20
        if bool(record.get("Clock In")) != bool(record.get("Clock Out")):
          score += 10
        return score

      def fetch_report_pages(session, endpoint, query, request_kwargs):
        """Fetch one BioTime report query and follow DRF pagination."""
        payloads = []
        next_url = endpoint
        next_params = dict(query)
        seen_urls = set()

        for _page in range(20):
          request_key = (next_url, tuple(sorted((next_params or {}).items())))
          if request_key in seen_urls:
            break
          seen_urls.add(request_key)

          try:
            response = session.get(
                next_url,
                params=next_params,
                timeout=20,
                **request_kwargs,
            )
          except requests.RequestException:
            return [], None

          if response.status_code != 200:
            return [], response.status_code

          try:
            payload = response.json()
          except ValueError:
            return [], response.status_code

          payloads.append(payload)

          if not isinstance(payload, dict):
            break
          next_link = payload.get("next")
          if not next_link:
            break
          next_link = str(next_link)
          if next_link.startswith("http://") or next_link.startswith("https://"):
            next_url = next_link
          elif next_link.startswith("/"):
            next_url = BASE_URL + next_link
          else:
            next_url = BASE_URL + "/" + next_link.lstrip("/")
          next_params = None

        return payloads, 200

      @strlit.cache_data(ttl=300, show_spinner=False)
      def fetch_biotime_calculated_range(
          start_date,
          end_date,
          employee_internal_id_map=None,
          expected_single_punch_keys=None,
          expected_attendance_keys=None,
      ):
        """Fetch BioTime's own calculated daily values for the whole range.

        The app does not require any BioTime report upload. It calls BioTime's
        /att/api/*Report/ endpoints directly and merges the best server-calculated
        value per employee/date. Single-punch records require Actual WT; normal
        IN/OUT records require Total WT. Missing server values fall back to the
        local minute-precision calculation already built from raw transactions.
        """
        start_text = start_date.strftime("%Y-%m-%d")
        end_text = end_date.strftime("%Y-%m-%d")
        employee_internal_id_map = employee_internal_id_map or {}
        expected_single_punch_keys = set(expected_single_punch_keys or ())
        expected_attendance_keys = set(expected_attendance_keys or ())
        token = get_auth_token()

        # dailyActivityReport normally exposes both Actual WT and Total WT. The
        # time-card reports are compatibility fallbacks and also preserve unusual
        # BioTime server calculations that cannot be reconstructed from raw punches.
        report_names = (
            "monthlyWorkHoursReport",
            "timeCardReport",
            "totalTimeCardReportV2",
            "dailyActivityReport",
            "monthlyPunchReport",
            "firstInLastOutReport",
        )
        query_variants = (
            {
                "start_date": start_text,
                "end_date": end_text,
                "page_size": 10000,
            },
            {
                "start_time": f"{start_text} 00:00:00",
                "end_time": f"{end_text} 23:59:59",
                "page_size": 10000,
            },
            {
                "from_date": start_text,
                "to_date": end_text,
                "page_size": 10000,
            },
        )

        # Reports work with HTTP Basic auth on BioTime even when other API routes
        # are license-gated. Token auth remains a compatibility fallback.
        auth_attempts = [
            {
                "auth": (EMAIL, PASSWORD),
                "headers": {"Accept": "application/json"},
            }
        ]
        if token:
          auth_attempts.append(
              {
                  "headers": {
                      "Authorization": f"Token {token}",
                      "Accept": "application/json",
                  }
              }
          )

        session = requests.Session()
        best_records = {}
        best_scores = {}

        def required_value_present(key, record):
          if key in expected_single_punch_keys:
            return bool(str(record.get("Actual WT", "") or "").strip())
          return bool(str(record.get("Report Total WT", "") or "").strip())

        def score_record(key, record, report_index):
          actual = str(record.get("Actual WT", "") or "").strip()
          total = str(record.get("Report Total WT", "") or "").strip()
          clock_in = str(record.get("Clock In", "") or "").strip()
          clock_out = str(record.get("Clock Out", "") or "").strip()
          is_single = key in expected_single_punch_keys

          # The value required by our agreed rule dominates the score. Prefer the
          # earlier/more specific report only when coverage is otherwise equal.
          score = 0
          if is_single:
            score += 1000 if actual else 0
            score += 100 if bool(clock_in) != bool(clock_out) else 0
            score += 20 if total else 0
          else:
            score += 1000 if total else 0
            score += 100 if clock_in and clock_out else 0
            score += 20 if actual else 0
          score += max(0, 20 - report_index)
          return score

        for report_index, report_name in enumerate(report_names):
          endpoint = f"{BASE_URL}/att/api/{report_name}/"
          endpoint_missing = False
          report_had_rows = False

          for query in query_variants:
            if endpoint_missing or report_had_rows:
              break

            for authorization in auth_attempts:
              payloads, status_code = fetch_report_pages(
                  session, endpoint, query, authorization
              )
              if status_code == 404:
                endpoint_missing = True
                break
              if not payloads:
                continue

              parsed_any = False
              for payload in payloads:
                for report_row in extract_report_rows(payload):
                  normalized_record = normalize_calculated_report_row(
                      report_row,
                      start_date,
                      end_date,
                      employee_internal_id_map,
                  )
                  if not normalized_record:
                    continue

                  normalized_record["Report Name"] = report_name

                  actual_wt = str(
                      normalized_record.get("Actual WT", "") or ""
                  ).strip()
                  total_wt = str(
                      normalized_record.get("Report Total WT", "") or ""
                  ).strip()
                  if not actual_wt and not total_wt:
                    continue

                  attendance_date = normalized_record["Attendance Date"]
                  employee_id = normalized_record["Employee ID"]
                  key = (attendance_date, employee_id)
                  score = score_record(key, normalized_record, report_index)
                  if score > best_scores.get(key, -1):
                    best_scores[key] = score
                    best_records[key] = normalized_record
                  parsed_any = True

              if parsed_any:
                report_had_rows = True
                break

          # Stop as soon as every known attendance day has the exact type of
          # server-calculated value required by the business rule. This keeps the
          # common path fast while still filling gaps from fallback reports.
          if expected_attendance_keys and all(
              key in best_records and required_value_present(key, best_records[key])
              for key in expected_attendance_keys
          ):
            break

        result = {}
        for (attendance_date, employee_id), record in best_records.items():
          result.setdefault(attendance_date, {})[employee_id] = record
        return result

      def fetch_paginated_api(session, path, params, headers, timeout=20):
        """Fetch a normal BioTime list endpoint, following pagination."""
        rows = []
        next_url = f"{BASE_URL}{path}"
        next_params = dict(params or {})
        seen_urls = set()

        for page_number in range(1, 50):
          request_key = (next_url, tuple(sorted((next_params or {}).items())))
          if request_key in seen_urls:
            break
          seen_urls.add(request_key)

          try:
            response = session.get(
                next_url,
                params=next_params,
                headers=headers,
                timeout=timeout,
            )
          except requests.RequestException:
            break

          if response.status_code != 200:
            break
          try:
            payload = response.json()
          except ValueError:
            break

          if isinstance(payload, list):
            rows.extend(item for item in payload if isinstance(item, dict))
            break
          if not isinstance(payload, dict):
            break

          page_rows = payload.get("data")
          if not isinstance(page_rows, list):
            page_rows = payload.get("results")
          if not isinstance(page_rows, list):
            page_rows = []
          rows.extend(item for item in page_rows if isinstance(item, dict))

          next_link = payload.get("next")
          if next_link:
            next_link = str(next_link)
            if next_link.startswith("http://") or next_link.startswith("https://"):
              next_url = next_link
            elif next_link.startswith("/"):
              next_url = BASE_URL + next_link
            else:
              next_url = BASE_URL + "/" + next_link.lstrip("/")
            next_params = None
            continue

          count = payload.get("count")
          try:
            count = int(count)
          except (TypeError, ValueError):
            count = len(rows)

          if len(rows) >= count or not page_rows:
            break

          # Some BioTime builds omit `next`; fall back to explicit page numbers.
          next_url = f"{BASE_URL}{path}"
          next_params = dict(params or {})
          next_params["page"] = page_number + 1

        return rows

      @strlit.cache_data(ttl=300, show_spinner=False)
      def load_monthly_attendance_bulk(start_date, end_date, requested_dates=()):
        """Load BioTime once for the Excel range, but build rows only for Excel dates.

        `requested_dates` is the exact set of date headers physically present in the
        uploaded attendance sheet. Data can be fetched in one range for speed, but the
        app never creates attendance output for dates that are not in the Excel file.
        """
        requested_date_set = {
            attendance_date
            for attendance_date in requested_dates
            if start_date <= attendance_date <= end_date
        }
        if not requested_date_set:
          cursor = start_date
          while cursor <= end_date:
            requested_date_set.add(cursor)
            cursor += timedelta(days=1)

        token = get_auth_token()
        if not token:
          raise RuntimeError("تعذر المصادقة مع BioTime")

        headers = {
            "Authorization": f"Token {token}",
            "Content-Type": "application/json",
        }
        session = requests.Session()

        devices = []
        for endpoint in ("/iclock/api/terminals/", "/iclock/api/devices/"):
          devices = fetch_paginated_api(
              session,
              endpoint,
              {"page_size": 1000},
              headers,
              timeout=15,
          )
          if devices:
            break

        terminal_map = {
            str(device.get("sn", "")): (
                device.get("alias")
                or device.get("terminal_name")
                or str(device.get("sn", ""))
            )
            for device in devices
            if device.get("sn")
        }

        all_employees = fetch_paginated_api(
            session,
            "/personnel/api/employees/",
            {"page_size": 1000},
            headers,
            timeout=20,
        )

        active_employees = {}
        employee_internal_id_map = {}

        for emp in all_employees:
          raw_code = str(emp.get("emp_code", "")).strip()
          cleaned_code = normalize_id(raw_code)

          is_active = (
              str(emp.get("is_active", True)).lower() in ("true", "1", "yes")
          )
          emp_status = str(emp.get("status", "0")).upper()
          enable_att = (
              str(emp.get("enable_attendance", True)).lower()
              in ("true", "1", "yes")
          )
          if not is_active or emp_status in ("1", "2", "D") or not enable_att:
            continue
          if not cleaned_code or cleaned_code in EXCLUDED_MANAGEMENT_CODES:
            continue

          first_name = str(emp.get("first_name", "") or "").strip()
          last_name = str(emp.get("last_name", "") or "").strip()
          if first_name.lower() == "none":
            first_name = ""
          if last_name.lower() == "none":
            last_name = ""
          full_name = f"{first_name} {last_name}".strip()

          dept_data = emp.get("department", {})
          dept_name = (
              dept_data.get("dept_name")
              if isinstance(dept_data, dict)
              else str(emp.get("department", "") or "")
          )
          if not dept_name or str(dept_name).lower() == "none":
            dept_name = "غير محدد"

          active_employees[cleaned_code] = {
              "name": clean_txt(full_name or f"موظف {cleaned_code}"),
              "dept": clean_txt(dept_name),
          }

          internal_id = normalize_id(emp.get("id"))
          if internal_id:
            employee_internal_id_map[internal_id] = cleaned_code

        employee_catalog = {
            employee_id: employee_data["name"]
            for employee_id, employee_data in active_employees.items()
        }

        leave_records = fetch_paginated_api(
            session,
            "/att/api/leave/",
            {"page_size": 5000},
            headers,
            timeout=15,
        )
        if not leave_records:
          leave_records = fetch_paginated_api(
              session,
              "/iclock/api/leave/",
              {"page_size": 5000},
              headers,
              timeout=15,
          )

        leave_by_date = {}
        for leave in leave_records:
          raw_code = (
              leave.get("emp_code")
              or leave.get("employee_code")
              or leave.get("employee_emp_code")
              or ""
          )
          cleaned_code = normalize_id(raw_code)
          if not cleaned_code:
            internal_employee = normalize_id(
                leave.get("employee_id") or leave.get("employee") or leave.get("emp")
            )
            cleaned_code = employee_internal_id_map.get(internal_employee, "")
          if cleaned_code not in active_employees:
            continue

          start_value = (
              leave.get("start_time")
              or leave.get("start_date")
              or leave.get("start_datetime")
          )
          end_value = (
              leave.get("end_time")
              or leave.get("end_date")
              or leave.get("end_datetime")
          )
          if not start_value or not end_value:
            continue

          try:
            leave_start = datetime.strptime(str(start_value)[:10], "%Y-%m-%d").date()
            leave_end = datetime.strptime(str(end_value)[:10], "%Y-%m-%d").date()
          except ValueError:
            continue

          leave_name = "إجازة"
          leave_type = leave.get("leave_type")
          if isinstance(leave_type, dict):
            leave_name = leave_type.get("leave_name") or leave_name
          elif leave_type not in (None, ""):
            leave_name = str(leave_type)
          elif leave.get("leave_name"):
            leave_name = str(leave.get("leave_name"))

          cursor = max(start_date, leave_start)
          last_date = min(end_date, leave_end)
          while cursor <= last_date:
            leave_by_date.setdefault(cursor, {})[cleaned_code] = clean_txt(leave_name)
            cursor += timedelta(days=1)

        raw_logs = fetch_paginated_api(
            session,
            "/iclock/api/transactions/",
            {
                "start_time": start_date.strftime("%Y-%m-%d") + " 00:00:00",
                "end_time": (end_date + timedelta(days=1)).strftime("%Y-%m-%d")
                + " 05:00:00",
                "page_size": 5000,
            },
            headers,
            timeout=30,
        )

        punches_by_emp_date = {}
        for log in raw_logs:
          cleaned_code = normalize_id(log.get("emp_code"))
          if not cleaned_code:
            internal_emp = normalize_id(log.get("emp"))
            cleaned_code = employee_internal_id_map.get(internal_emp, "")
          if cleaned_code not in active_employees:
            continue

          punch_value = log.get("punch_time")
          if not punch_value:
            continue
          try:
            punch_time = datetime.strptime(
                str(punch_value)[:19],
                "%Y-%m-%d %H:%M:%S",
            )
          except ValueError:
            continue

          device_sn = str(log.get("terminal_sn", "") or "")
          device_name = (
              log.get("terminal_alias")
              or log.get("terminal_name")
              or terminal_map.get(device_sn, device_sn or "جهاز رئيسي")
          )
          punch_kind = classify_transaction_punch(log, punch_time)
          punches_by_emp_date.setdefault(
              (cleaned_code, punch_time.date()), []
          ).append(
              {
                  "time": punch_time,
                  "device": device_name,
                  "kind": punch_kind,
              }
          )

        # Remove duplicate punches within 60 seconds once for the whole month.
        # Keep the punch-state classification because it is essential for avoiding
        # false pairings of two IN punches or two OUT punches.
        for key, punches in list(punches_by_emp_date.items()):
          punches = sorted(punches, key=lambda item: item["time"])
          filtered = []
          for punch in punches:
            if (
                not filtered
                or abs(
                    (punch["time"] - filtered[-1]["time"]).total_seconds()
                ) > 60
                or punch["kind"] != filtered[-1]["kind"]
            ):
              filtered.append(punch)
          punches_by_emp_date[key] = filtered

        all_attendance = {}
        today = datetime.now(SYRIA_TZ).date()

        for cursor in sorted(requested_date_set):
          date_map = {}
          leave_map = leave_by_date.get(cursor, {})

          for code, emp_data in active_employees.items():
            name = emp_data["name"]
            dept = emp_data["dept"]

            day_punches = [
                punch
                for punch in punches_by_emp_date.get((code, cursor), [])
                if punch["time"].hour >= 5
            ]
            next_morning = [
                punch
                for punch in punches_by_emp_date.get(
                    (code, cursor + timedelta(days=1)), []
                )
                if punch["time"].hour < 5
            ]

            if not day_punches and not next_morning:
              if code in leave_map:
                date_map[code] = {
                    "Employee ID": code,
                    "First Name": name,
                    "Department": dept,
                    "Date": cursor.strftime("%Y-%m-%d"),
                    "Clock In": "",
                    "Clock Out": "",
                    "Actual WT": "",
                    "Calculated WT": "",
                    "Total WT": "",
                    "Status": f"Leave - {leave_map[code]}",
                }
              else:
                date_map[code] = {
                    "Employee ID": code,
                    "First Name": name,
                    "Department": dept,
                    "Date": cursor.strftime("%Y-%m-%d"),
                    "Clock In": "",
                    "Clock Out": "",
                    "Actual WT": "",
                    "Calculated WT": "",
                    "Total WT": "",
                    "Status": "Absence(A)",
                }
              continue

            all_candidates = sorted(
                day_punches + next_morning,
                key=lambda punch: punch["time"],
            )
            punch_count_for_day = len(all_candidates)
            odd_even_multi_punch = (
                punch_count_for_day > 2 and punch_count_for_day % 2 == 0
            )
            odd_even_total_wt = (
                calculate_odd_even_punch_total(
                    [punch["time"] for punch in all_candidates]
                )
                if odd_even_multi_punch
                else ""
            )

            explicit_in = [p for p in day_punches if p["kind"] == "in"]
            explicit_out = [p for p in all_candidates if p["kind"] == "out"]
            # BioTime's day-change rule associates punches before 05:00 with the
            # previous work day. Treat those as OUT candidates even if a terminal
            # labelled the punch ambiguously.
            for next_punch in next_morning:
              if next_punch not in explicit_out:
                explicit_out.append(next_punch)
            unknown_day = [p for p in day_punches if p["kind"] == "unknown"]
            unknown_next = [p for p in next_morning if p["kind"] == "unknown"]

            # When a transaction carries a real punch state, trust it. Unknown
            # punches are used only to fill a side that BioTime did not label.
            in_candidates = list(explicit_in)
            out_candidates = list(explicit_out)

            for punch in unknown_day:
              if punch["time"].hour < SINGLE_PUNCH_OUT_HOUR:
                if not explicit_in:
                  in_candidates.append(punch)
              else:
                if not explicit_out:
                  out_candidates.append(punch)

            if unknown_next and not explicit_out:
              out_candidates.extend(unknown_next)

            # Final fallback for devices that send no punch state at all: classify
            # all morning punches as IN candidates and all afternoon/evening punches
            # as OUT candidates. Never pair two morning-only punches together.
            if not in_candidates and not out_candidates:
              in_candidates = [
                  p for p in day_punches
                  if p["time"].hour < SINGLE_PUNCH_OUT_HOUR
              ]
              out_candidates = [
                  p for p in day_punches
                  if p["time"].hour >= SINGLE_PUNCH_OUT_HOUR
              ] + list(next_morning)

            if odd_even_multi_punch and odd_even_total_wt:
              # User-adjusted BioTime punches follow sequential odd/even pairing:
              # 1->2, 3->4, ... . Keep first/last only for display; the duty value
              # is the SUM of every completed pair, not first-to-last.
              clock_in_punch = all_candidates[0]
              clock_out_punch = all_candidates[-1]
            else:
              clock_in_punch = (
                  min(in_candidates, key=lambda p: p["time"])
                  if in_candidates
                  else None
              )
              clock_out_punch = (
                  max(out_candidates, key=lambda p: p["time"])
                  if out_candidates
                  else None
              )

            clock_in_dt = clock_in_punch["time"] if clock_in_punch else None
            clock_out_dt = clock_out_punch["time"] if clock_out_punch else None

            # If both sides somehow resolve in reverse order on the same work date,
            # keep the side that is credible instead of manufacturing a tiny shift.
            if (
                clock_in_dt is not None
                and clock_out_dt is not None
                and clock_out_dt <= clock_in_dt
                and clock_out_dt.date() == clock_in_dt.date()
            ):
              if clock_in_dt.hour < SINGLE_PUNCH_OUT_HOUR <= clock_out_dt.hour:
                pass
              elif clock_in_dt.hour < SINGLE_PUNCH_OUT_HOUR:
                clock_out_dt = None
                clock_out_punch = None
              else:
                clock_in_dt = None
                clock_in_punch = None

            is_single_punch = bool(clock_in_dt) != bool(clock_out_dt)
            if odd_even_multi_punch and odd_even_total_wt:
              selected_work_time = odd_even_total_wt
            else:
              selected_work_time = calculate_monthly_working_hours(
                  cursor,
                  clock_in_dt,
                  clock_out_dt,
              )
            actual_wt = selected_work_time if is_single_punch else ""
            total_wt = selected_work_time if not is_single_punch else ""

            # A transaction exists, so this is attendance even if only one side is
            # present. Actual WT is schedule-aware and never raw first-last duration.
            is_late = bool(
                clock_in_dt
                and (
                    clock_in_dt.hour > 9
                    or (clock_in_dt.hour == 9 and clock_in_dt.minute > 15)
                )
            )

            if clock_in_dt and not clock_out_dt:
              status = "Late(LT) / Missing OUT" if is_late else "Present(P) / Missing OUT"
            elif clock_out_dt and not clock_in_dt:
              status = "Present(P) / Missing IN"
            else:
              status = "Late(LT)" if is_late else "Present(P)"

            date_map[code] = {
                "Employee ID": code,
                "First Name": name,
                "Department": dept,
                "Date": cursor.strftime("%Y-%m-%d"),
                "Clock In": clock_in_dt.strftime("%H:%M") if clock_in_dt else "",
                "Clock Out": clock_out_dt.strftime("%H:%M") if clock_out_dt else "",
                "Actual WT": actual_wt,
                "Calculated WT": selected_work_time,
                "Total WT": total_wt,
                "Punch Count": punch_count_for_day,
                "Odd Even Paired": odd_even_multi_punch,
                "Status": status,
            }

          all_attendance[cursor] = date_map

        return (
            all_attendance,
            employee_catalog,
            employee_internal_id_map,
            active_employees,
        )

      def detect_month_sheet(workbook):
        preferred = [
            sheet_name
            for sheet_name in workbook.sheetnames
            if any(
                key in sheet_name
                for key in [
                    "مجموع ساعات الدوام",
                    "دوام",
                    "ساعات",
                    "حضور",
                    "August",
                    "اغسطس",
                    "أغسطس",
                ]
            )
        ]
        return workbook[
            preferred[0] if preferred else workbook.sheetnames[0]
        ]

      def detect_date_columns(ws):
        """The attendance template has its dates in row 1."""
        date_columns = {}

        for column in range(1, ws.max_column + 1):
          value = ws.cell(row=1, column=column).value
          parsed = None

          if isinstance(value, datetime):
            parsed = value.date()
          elif hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
            parsed = value
          elif isinstance(value, str):
            raw = value.strip()
            for fmt in (
                "%Y-%m-%d",
                "%d/%m/%Y",
                "%d-%m-%Y",
                "%d/%m/%y",
                "%d-%m-%y",
            ):
              try:
                parsed = datetime.strptime(raw, fmt).date()
                break
              except ValueError:
                continue

          if parsed:
            date_columns[column] = parsed

        return date_columns

      def detect_employee_total_row(ws):
        """Return the summary/totals row that ends the employee roster.

        The Golden Palace template normally labels this row `الإجمالي الكلي`. If the
        label changes, use an Excel table totals row as a safe fallback.
        """
        for row_number in range(2, ws.max_row + 1):
          label = clean_txt(ws.cell(row=row_number, column=3).value)
          if label and ("الإجمالي" in label or "المجموع" in label):
            return row_number

        for table_name in ws.tables:
          try:
            table = ws.tables[table_name]
            if int(table.totalsRowCount or 0) > 0:
              match = re.match(r"^[A-Z]+(\d+):[A-Z]+(\d+)$", str(table.ref))
              if match:
                return int(match.group(2))
          except Exception:
            continue

        return ws.max_row + 1

      def find_available_employee_rows(ws, total_row):
        """Find reserved blank staff rows before totals without overwriting anyone.

        A row is available only when BOTH BioTime ID (column B) and employee name
        (column C) are empty. Existing named rows are never reused, even when B is blank.
        """
        rows = []
        for row_number in range(2, total_row):
          biotime_id = normalize_id(ws.cell(row=row_number, column=2).value)
          employee_name = clean_txt(ws.cell(row=row_number, column=3).value)
          if not biotime_id and not employee_name:
            rows.append(row_number)
        return rows

      def excel_employee_id_value(employee_id):
        """Write numeric BioTime IDs as numbers while preserving non-numeric IDs."""
        normalized = normalize_id(employee_id)
        return int(normalized) if normalized.isdigit() else normalized

      def time_to_excel_value(value):
        """Convert HH:MM to a real Excel time fraction."""
        if value is None or str(value).strip() == "":
          return None

        raw = str(value).strip()
        parts = raw.split(":")
        if len(parts) < 2:
          return None

        try:
          hours = int(parts[0])
          minutes = int(parts[1])
          seconds = int(parts[2]) if len(parts) > 2 else 0
          return (hours * 3600 + minutes * 60 + seconds) / 86400
        except (TypeError, ValueError):
          return None

      def excel_col_letter(column_number):
        return get_column_letter(column_number)

      def find_existing_duration_style_id(ws, date_columns):
        """Reuse an existing raw OOXML [h]:mm style id from the template."""
        # Never rebuild styles.xml. openpyxl may expose duplicate/generated style ids,
        # so resolve the cell style back to the FIRST style already stored in the
        # uploaded workbook. That index is the safe worksheet s= style id.
        workbook_styles = list(getattr(ws.parent, "_cell_styles", []))
        for row_number in range(2, ws.max_row + 1):
          for column_number in date_columns:
            cell = ws.cell(row=row_number, column=column_number)
            if str(cell.number_format or "").strip().lower() != "[h]:mm":
              continue
            for style_index, existing_style in enumerate(workbook_styles):
              if existing_style == cell._style:
                return int(style_index)
        return None

      def find_existing_cell_style_id(ws, row_number, column_number):
        """Resolve a cell's style to the original raw worksheet style index."""
        cell = ws.cell(row=row_number, column=column_number)
        workbook_styles = list(getattr(ws.parent, "_cell_styles", []))
        for style_index, existing_style in enumerate(workbook_styles):
          if existing_style == cell._style:
            return int(style_index)
        return None

      def _xml_cell_fragment(coordinate, value, attrs="", tag_prefix="", style_id=None):
        """Build one worksheet cell while preserving the sheet namespace prefix."""
        attrs = attrs or f' r="{coordinate}"'
        if not re.search(r'\br="[^"]+"', attrs):
          attrs = f' r="{coordinate}"' + attrs
        attrs = re.sub(r'\s+t="[^"]*"', '', attrs)
        if style_id is not None:
          attrs = re.sub(r'\s+s="\d+"', '', attrs)
          attrs += f' s="{int(style_id)}"'

        cell_tag = f"{tag_prefix}c"
        value_tag = f"{tag_prefix}v"
        is_tag = f"{tag_prefix}is"
        text_tag = f"{tag_prefix}t"

        if value is None or value == "":
          return f"<{cell_tag}{attrs}/>"
        if isinstance(value, str):
          safe_text = (
              value.replace("&", "&amp;").replace("<", "&lt;")
                   .replace(">", "&gt;").replace('"', "&quot;")
          )
          return (
              f'<{cell_tag}{attrs} t="inlineStr"><{is_tag}><{text_tag}>'
              f'{safe_text}</{text_tag}></{is_tag}></{cell_tag}>'
          )
        return f"<{cell_tag}{attrs}><{value_tag}>{value}</{value_tag}></{cell_tag}>"

      def patch_cell_value_xml(xml_text, row_number, column_number, value, style_id=None):
        """Overwrite/create one worksheet cell without rebuilding the workbook package."""
        coordinate = f"{get_column_letter(column_number)}{row_number}"
        cell_tag_pattern = r'(?:[A-Za-z_][\w.\-]*:)?c'
        pattern = re.compile(
            rf'<(?P<tag>{cell_tag_pattern})\b(?=[^>]*\br="{re.escape(coordinate)}")'
            rf'(?P<attrs>[^>]*)/>'
            rf'|<(?P<tag2>{cell_tag_pattern})\b(?=[^>]*\br="{re.escape(coordinate)}")'
            rf'(?P<attrs2>[^>]*)>.*?</(?P=tag2)>',
            re.DOTALL,
        )
        match = pattern.search(xml_text)
        if match:
          tag_name = match.group("tag") or match.group("tag2") or "c"
          tag_prefix = tag_name[:-1]
          attrs = match.group("attrs") if match.group("attrs") is not None else match.group("attrs2")
          replacement = _xml_cell_fragment(
              coordinate, value, attrs=attrs, tag_prefix=tag_prefix, style_id=style_id
          )
          return xml_text[:match.start()] + replacement + xml_text[match.end():]

        # Some templates omit physically empty cells from the worksheet XML. Create
        # the cell inside its existing row, preserving the row's namespace prefix and
        # inserting it before the next higher column when possible.
        row_tag_pattern = r'(?:[A-Za-z_][\w.\-]*:)?row'
        row_pattern = re.compile(
            rf'<(?P<rowtag>{row_tag_pattern})\b(?=[^>]*\br="{row_number}")'
            rf'(?P<rowattrs>[^>]*)>(?P<body>.*?)</(?P=rowtag)>',
            re.DOTALL,
        )
        row_match = row_pattern.search(xml_text)
        if not row_match:
          raise RuntimeError(f"Excel row {row_number} was not found in worksheet XML")

        row_tag = row_match.group("rowtag")
        tag_prefix = row_tag[:-3]
        body = row_match.group("body")
        new_cell = _xml_cell_fragment(
            coordinate, value, tag_prefix=tag_prefix, style_id=style_id
        )

        target_column = column_number
        insert_at = len(body)
        existing_cell_pattern = re.compile(
            rf'<(?:[A-Za-z_][\w.\-]*:)?c\b[^>]*\br="([A-Z]+){row_number}"',
            re.DOTALL,
        )
        for existing_match in existing_cell_pattern.finditer(body):
          letters = existing_match.group(1)
          existing_column = 0
          for letter in letters:
            existing_column = existing_column * 26 + (ord(letter) - 64)
          if existing_column > target_column:
            insert_at = existing_match.start()
            break

        new_body = body[:insert_at] + new_cell + body[insert_at:]
        replacement_row = (
            f'<{row_tag}{row_match.group("rowattrs")}>{new_body}</{row_tag}>'
        )
        return (
            xml_text[:row_match.start()] + replacement_row + xml_text[row_match.end():]
        )

      def patch_cell_formula_xml(xml_text, row_number, column_number, formula):
        """Replace a formula and remove its stale cached result, with namespace support."""
        coordinate = f"{get_column_letter(column_number)}{row_number}"
        cell_tag_pattern = r'(?:[A-Za-z_][\w.\-]*:)?c'
        pattern = re.compile(
            rf'<(?P<tag>{cell_tag_pattern})\b(?=[^>]*\br="{re.escape(coordinate)}")'
            rf'(?P<attrs>[^>]*)>(?P<body>.*?)</(?P=tag)>',
            re.DOTALL,
        )
        match = pattern.search(xml_text)
        if not match:
          raise RuntimeError(f"Excel formula cell {coordinate} was not found")
        tag_name = match.group("tag")
        tag_prefix = tag_name[:-1]
        attrs = match.group("attrs")
        body = match.group("body")
        formula_text = str(formula or "")
        if formula_text.startswith("="):
          formula_text = formula_text[1:]
        safe_formula = formula_text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        escaped_prefix = re.escape(tag_prefix)
        formula_pattern = rf'<{escaped_prefix}f(?:\s[^>]*)?>.*?</{escaped_prefix}f>'
        value_pattern = rf'<{escaped_prefix}v>.*?</{escaped_prefix}v>'
        if re.search(formula_pattern, body, re.DOTALL):
          body = re.sub(
              formula_pattern,
              f'<{tag_prefix}f>{safe_formula}</{tag_prefix}f>',
              body, count=1, flags=re.DOTALL,
          )
        else:
          body = f'<{tag_prefix}f>{safe_formula}</{tag_prefix}f>' + body
        body = re.sub(value_pattern, '', body, flags=re.DOTALL)
        replacement = f"<{tag_name}{attrs}>{body}</{tag_name}>"
        return xml_text[:match.start()] + replacement + xml_text[match.end():]

      def build_total_formula_updates(ws, employee_matches, date_columns, today_date):
        """Keep today's attendance visible, but exclude today/future dates from totals."""
        prior_dates = [
            (column, attendance_date)
            for column, attendance_date in date_columns.items()
            if attendance_date < today_date
        ]
        if not prior_dates:
          return []
        cutoff_column, cutoff_date = max(prior_dates, key=lambda item: item[1])
        cutoff_raw = ws.cell(row=1, column=cutoff_column).value
        if isinstance(cutoff_raw, datetime) or hasattr(cutoff_raw, "strftime"):
          cutoff_header = cutoff_raw.strftime("%d/%m/%Y")
        else:
          cutoff_header = str(cutoff_raw).strip() or cutoff_date.strftime("%d/%m/%Y")

        end_pattern = re.compile(r':\[([0-3]?\d/[01]?\d/\d{4})\]')
        updates = []
        for match in employee_matches:
          row_number = match["row"]
          for column in range(1, ws.max_column + 1):
            raw_formula = ws.cell(row=row_number, column=column).value
            if not (isinstance(raw_formula, str) and raw_formula.startswith("=")):
              continue
            changed = False
            def replace_end_date(regex_match):
              nonlocal changed
              raw_date = regex_match.group(1)
              try:
                parsed_date = datetime.strptime(raw_date, "%d/%m/%Y").date()
              except ValueError:
                return regex_match.group(0)
              if parsed_date >= today_date:
                changed = True
                return f":[{cutoff_header}]"
              return regex_match.group(0)
            new_formula = end_pattern.sub(replace_end_date, raw_formula)
            if changed and new_formula != raw_formula:
              updates.append({"row": row_number, "column": column, "formula": new_formula})
        return updates

      def find_sheet_xml_path(zip_file, sheet_name):
        """Resolve a workbook sheet name to its worksheet XML path."""
        main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        rel_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
        pkg_rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"

        workbook_xml = zip_file.read("xl/workbook.xml").decode("utf-8")
        rels_xml = zip_file.read("xl/_rels/workbook.xml.rels").decode("utf-8")

        workbook_root = ET.fromstring(workbook_xml)
        rel_root = ET.fromstring(rels_xml)

        rel_targets = {}
        for rel in rel_root.findall(f"{{{pkg_rel_ns}}}Relationship"):
          rel_targets[rel.get("Id")] = rel.get("Target")

        for sheet in workbook_root.findall(f"{{{main_ns}}}sheets/{{{main_ns}}}sheet"):
          if sheet.get("name") == sheet_name:
            rel_id = sheet.get(f"{{{rel_ns}}}id")
            target = rel_targets.get(rel_id)
            if not target:
              raise RuntimeError(f"Could not resolve worksheet relationship for {sheet_name}")
            target = target.lstrip("/")
            if not target.startswith("xl/"):
              target = "xl/" + target
            return target

        raise RuntimeError(f"Worksheet not found: {sheet_name}")

      def export_template_preserving_package(
          original_bytes, sheet_name, cell_updates, formula_updates=None, duration_style_id=None
      ):
        """Patch values/formulas into the original OOXML package without rebuilding it.

        IMPORTANT: xl/styles.xml is copied byte-for-byte. Numeric attendance cells may
        reuse an EXISTING [h]:mm style id from the uploaded template, but no style
        definitions or workbook metadata are regenerated.
        """
        input_buffer = io.BytesIO(original_bytes)
        output_buffer = io.BytesIO()
        formula_updates = formula_updates or []
        with zipfile.ZipFile(input_buffer, "r") as zin:
          sheet_xml_path = find_sheet_xml_path(zin, sheet_name)
          sheet_xml_text = zin.read(sheet_xml_path).decode("utf-8")

          for update in cell_updates:
            value = update["value"]
            explicit_style_id = update.get("style_id")
            # Approved BioTime ID/name cells may provide their existing row style.
            # Otherwise only real numeric attendance durations use [h]:mm.
            use_cell_style = explicit_style_id
            if use_cell_style is None:
              use_cell_style = (
                  duration_style_id
                  if duration_style_id is not None
                  and value not in (None, "")
                  and not isinstance(value, str)
                  else None
              )
            sheet_xml_text = patch_cell_value_xml(
                sheet_xml_text,
                update["row"], update["column"], value,
                style_id=use_cell_style,
            )

          for formula_update in formula_updates:
            sheet_xml_text = patch_cell_formula_xml(
                sheet_xml_text,
                formula_update["row"], formula_update["column"], formula_update["formula"],
            )
          patched_sheet_xml = sheet_xml_text.encode("utf-8")

          workbook_xml_text = zin.read("xl/workbook.xml").decode("utf-8")
          calc_pattern = re.compile(r"<calcPr\b([^>]*)/>")
          calc_match = calc_pattern.search(workbook_xml_text)
          if calc_match:
            calc_attrs = re.sub(
                r'\s+(?:calcMode|calcOnSave|fullCalcOnLoad|forceFullCalc)="[^"]*"',
                "", calc_match.group(1),
            )
            calc_replacement = (
                f'<calcPr{calc_attrs} calcMode="auto" calcOnSave="1" '
                'fullCalcOnLoad="1" forceFullCalc="1"/>'
            )
            workbook_xml_text = (
                workbook_xml_text[:calc_match.start()] + calc_replacement
                + workbook_xml_text[calc_match.end():]
            )
          else:
            workbook_xml_text = workbook_xml_text.replace(
                "</workbook>",
                '<calcPr calcMode="auto" calcOnSave="1" fullCalcOnLoad="1" forceFullCalc="1"/></workbook>',
            )
          patched_workbook_xml = workbook_xml_text.encode("utf-8")

          with zipfile.ZipFile(output_buffer, "w") as zout:
            for item in zin.infolist():
              data = zin.read(item.filename)
              if item.filename == sheet_xml_path:
                data = patched_sheet_xml
              elif item.filename == "xl/workbook.xml":
                data = patched_workbook_xml
              # Everything else, especially xl/styles.xml, is copied exactly.
              zout.writestr(item, data)
        output_buffer.seek(0)
        return output_buffer.getvalue()

      @strlit.cache_data(ttl=300, show_spinner=False)
      def fetch_month_date(attendance_date):
        date_str = attendance_date.strftime("%Y-%m-%d")
        is_date_today = attendance_date == now_syria.date()
        return load_attendance_data_from_api(
            date_str,
            attendance_date,
            is_date_today,
        )

      if uploaded_template is not None:
        monthly_loading_overlay = show_loading_overlay(
            "جاري تحميل ومعالجة بيانات الدوام..."
        )
        export_loading_overlay = None
        try:
          # Keep the exact original XLSX/XLSM bytes. The final export will patch
          # only the attendance cells into this original package.
          original_template_bytes = uploaded_template.getvalue()
          keep_vba = uploaded_template.name.lower().endswith(".xlsm")

          template_wb = openpyxl.load_workbook(
              io.BytesIO(original_template_bytes),
              keep_vba=keep_vba,
              data_only=False,
          )
          ws_target = detect_month_sheet(template_wb)

          date_columns = detect_date_columns(ws_target)
          if not date_columns:
            strlit.error("لم يتم العثور على أعمدة التواريخ في الصف الأول من ملف الدوام.")
            raise RuntimeError("No monthly date columns detected")

          # Use the template's own existing [h]:mm attendance style.
          # This fixes General-formatted blank cells without modifying styles.xml.
          duration_style_id = find_existing_duration_style_id(ws_target, date_columns)

          # IMPORTANT:
          # Excel Column B is the BioTime ID.
          # Excel Column A is NOT used for employee matching.
          # Employee names are NOT used for automatic matching.
          # New BioTime staff are added only after explicit user approval.
          employee_total_row = detect_employee_total_row(ws_target)
          available_new_staff_rows = find_available_employee_rows(
              ws_target, employee_total_row
          )

          excel_employees = []
          for row in range(2, employee_total_row):
            biotime_id = normalize_id(ws_target.cell(row=row, column=2).value)
            employee_name = clean_txt(ws_target.cell(row=row, column=3).value)

            if not biotime_id and not employee_name:
              continue

            excel_employees.append(
                {
                    "row": row,
                    "biotime_id": biotime_id,
                    "name": employee_name,
                }
            )

          if not excel_employees:
            strlit.error(
                "لم يتم العثور على موظفين في الملف. تأكد من أن BioTime ID موجود في العمود B."
            )
            raise RuntimeError("No employee rows detected")

          # FAST MONTHLY LOAD:
          # Employees, devices, leave and transactions are fetched once for the
          # entire Excel date range. BioTime calculated report values are requested
          # once in bulk. Single-punch days use Actual WT; normal days use Total WT.
          # If the report endpoint is unavailable, local values use the exact
          # minute precision shown by BioTime Clock In / Clock Out.
          sorted_dates = sorted(set(date_columns.values()))
          dates_list = [
              (column, attendance_date)
              for column, attendance_date in date_columns.items()
              if attendance_date <= now_syria.date()
          ]

          if not dates_list:
            strlit.error("لا توجد تواريخ حالية قابلة للمعالجة في ملف الدوام.")
            raise RuntimeError("No current attendance dates to process")

          range_start = min(attendance_date for _column, attendance_date in dates_list)
          range_end = max(attendance_date for _column, attendance_date in dates_list)

          progress = strlit.progress(
              0,
              text="المرحلة 1/4 — تحميل بيانات BioTime الشهرية...",
          )

          requested_dates_for_load = tuple(
              sorted({attendance_date for _column, attendance_date in dates_list})
          )
          (
              all_attendance,
              employee_catalog,
              employee_internal_id_map,
              active_employee_details,
          ) = load_monthly_attendance_bulk(
              range_start,
              range_end,
              requested_dates=requested_dates_for_load,
          )

          # Find ACTIVE BioTime employees whose BioTime ID does not exist in Excel B.
          # This is an ID comparison only. Names/departments are shown for review but
          # are never used to auto-match an existing Excel employee.
          excel_biotime_ids = {
              employee["biotime_id"]
              for employee in excel_employees
              if employee["biotime_id"]
          }
          missing_active_ids = [
              employee_id
              for employee_id in active_employee_details
              if employee_id not in excel_biotime_ids
          ]
          missing_active_ids.sort(
              key=lambda value: (0, int(value)) if str(value).isdigit() else (1, str(value))
          )

          template_signature = hashlib.sha256(original_template_bytes).hexdigest()[:16]
          candidate_signature = hashlib.sha256(
              "|".join(missing_active_ids).encode("utf-8")
          ).hexdigest()[:10]
          approval_state_key = (
              f"approved_missing_biotime_staff_{template_signature}_{candidate_signature}"
          )
          approved_new_staff_ids = strlit.session_state.get(approval_state_key)

          if missing_active_ids and approved_new_staff_ids is None:
            # Stop before Excel modification. The user must explicitly review and approve.
            hide_loading_overlay(monthly_loading_overlay)
            monthly_loading_overlay = None
            progress.empty()

            strlit.markdown(
                '<div class="gp-section-title">👤 موظفون موجودون في BioTime وغير موجودين في Excel</div>'
                '<div class="gp-file-note">راجع القائمة واختر فقط من تريد إضافته إلى ملف Excel. '
                'لن يتم إضافة أي موظف بدون موافقتك، والمطابقة هنا حسب BioTime ID فقط.</div>',
                unsafe_allow_html=True,
            )

            review_rows = []
            for employee_id in missing_active_ids:
              employee_info = active_employee_details.get(employee_id, {})
              review_rows.append(
                  {
                      "BioTime ID": employee_id,
                      "الاسم": employee_info.get("name", ""),
                      "القسم": employee_info.get("dept", "غير محدد"),
                  }
              )
            strlit.dataframe(
                pd.DataFrame(review_rows),
                use_container_width=True,
                hide_index=True,
            )

            strlit.caption(
                f"صفوف Excel الفارغة المتاحة للإضافة: {len(available_new_staff_rows)}"
            )

            def missing_staff_option_label(employee_id):
              employee_info = active_employee_details.get(employee_id, {})
              return (
                  f"{employee_info.get('name', '')} — BioTime ID {employee_id}"
                  f" — {employee_info.get('dept', 'غير محدد')}"
              )

            selected_new_staff = strlit.multiselect(
                "اختر الموظفين الذين توافق على إضافتهم",
                options=missing_active_ids,
                default=[],
                format_func=missing_staff_option_label,
                key=f"missing_staff_review_{template_signature}_{candidate_signature}",
            )

            too_many_selected = len(selected_new_staff) > len(available_new_staff_rows)
            if too_many_selected:
              strlit.error(
                  "عدد الموظفين المختارين أكبر من عدد الصفوف الفارغة المتاحة في ملف Excel. "
                  "خفّض الاختيار أو أضف صفوف موظفين فارغة إلى القالب أولاً."
              )

            if strlit.button(
                "✅ تأكيد الاختيار ومتابعة معالجة الملف",
                use_container_width=True,
                disabled=too_many_selected,
                key=f"confirm_missing_staff_{template_signature}_{candidate_signature}",
            ):
              strlit.session_state[approval_state_key] = tuple(selected_new_staff)
              strlit.rerun()

            strlit.stop()

          approved_new_staff_ids = [
              normalize_id(employee_id)
              for employee_id in (approved_new_staff_ids or ())
              if normalize_id(employee_id) in active_employee_details
              and normalize_id(employee_id) in missing_active_ids
          ]
          if len(approved_new_staff_ids) > len(available_new_staff_rows):
            raise RuntimeError(
                "The approved BioTime staff exceed the empty employee rows available in Excel."
            )

          approved_new_staff_rows = []
          for row_number, employee_id in zip(
              available_new_staff_rows, approved_new_staff_ids
          ):
            employee_info = active_employee_details[employee_id]
            approved_new_staff_rows.append(
                {
                    "row": row_number,
                    "employee_id": employee_id,
                    "name": employee_info.get("name", f"موظف {employee_id}"),
                    "dept": employee_info.get("dept", "غير محدد"),
                }
            )

          target_employee_ids = (
              set(excel_biotime_ids) | set(approved_new_staff_ids)
          ) & set(active_employee_details.keys())

          progress.progress(
              0.62,
              text="المرحلة 2/4 — تحميل قيم Actual WT / Total WT من BioTime...",
          )

          # Tell the report fetcher exactly which type of value each employee/day
          # needs. This lets it stop early when BioTime coverage is complete while
          # still trying fallback report endpoints only for missing values.
          expected_attendance_keys = set()
          expected_single_punch_keys = set()
          for attendance_date, date_map in all_attendance.items():
            for employee_id, attendance in date_map.items():
              if employee_id not in target_employee_ids:
                continue
              status = str(attendance.get("Status", "") or "")
              if "Leave" in status or "Absence" in status:
                continue
              key = (attendance_date, employee_id)
              expected_attendance_keys.add(key)
              clock_in = str(attendance.get("Clock In", "") or "").strip()
              clock_out = str(attendance.get("Clock Out", "") or "").strip()
              if bool(clock_in) != bool(clock_out):
                expected_single_punch_keys.add(key)

          # Authoritative overlay: single-punch days use BioTime Actual WT; normal
          # attendance days use BioTime Total WT so uncapped/overtime hours remain visible.
          report_actual = fetch_biotime_calculated_range(
              range_start,
              range_end,
              employee_internal_id_map,
              expected_single_punch_keys,
              expected_attendance_keys,
          )
          report_actual_count = 0
          for attendance_date, report_date_map in report_actual.items():
            date_map = all_attendance.get(attendance_date, {})
            for employee_id, report_record in report_date_map.items():
              attendance = date_map.get(employee_id)
              if attendance is None:
                continue

              actual_wt = str(report_record.get("Actual WT", "") or "").strip()
              report_total_wt = str(
                  report_record.get("Report Total WT", "") or ""
              ).strip()

              report_clock_in = str(
                  report_record.get("Clock In", "") or ""
              ).strip()
              report_clock_out = str(
                  report_record.get("Clock Out", "") or ""
              ).strip()

              report_name = str(report_record.get("Report Name", "") or "")
              local_work_time = str(
                  attendance.get("Calculated WT", "") or ""
              ).strip()

              # Keep the raw-transaction punch shape when local data is complete.
              # A report endpoint may expose a different punch presentation even
              # though its hours are valid. Trust report clocks only for the exact
              # Monthly Worked Hrs source or when local transactions had no value.
              if (
                  report_clock_in or report_clock_out
              ) and (
                  report_name == "monthlyWorkHoursReport" or not local_work_time
              ):
                attendance["Clock In"] = report_clock_in
                attendance["Clock Out"] = report_clock_out

              final_clock_in = str(attendance.get("Clock In", "") or "").strip()
              final_clock_out = str(attendance.get("Clock Out", "") or "").strip()
              is_single_punch = bool(final_clock_in) != bool(final_clock_out)

              if is_single_punch:
                report_work_time = actual_wt or report_total_wt
              else:
                report_work_time = report_total_wt or actual_wt

              # For 4/6/8... punches, the user's BioTime setup is explicitly
              # adjusted as odd/even pairs (1->2, 3->4, ...). Keep the locally
              # calculated SUM of those pairs. Some report endpoints expose only
              # the first pair (for example 00:08) and must not overwrite it.
              if attendance.get("Odd Even Paired") and local_work_time:
                selected_work_time = local_work_time
              elif report_name == "monthlyWorkHoursReport" and report_work_time:
                selected_work_time = report_work_time
              else:
                selected_work_time = local_work_time or report_work_time

              if not selected_work_time:
                continue

              if is_single_punch:
                attendance["Actual WT"] = selected_work_time
              else:
                attendance["Total WT"] = selected_work_time
              attendance["Calculated WT"] = selected_work_time

              # A valid BioTime work-time row is attendance even if one punch is missing.
              if "Leave" not in str(attendance.get("Status", "")):
                clock_in = str(attendance.get("Clock In", "") or "").strip()
                clock_out = str(attendance.get("Clock Out", "") or "").strip()
                if clock_in and not clock_out:
                  attendance["Status"] = "Present(P) / Missing OUT"
                elif clock_out and not clock_in:
                  attendance["Status"] = "Present(P) / Missing IN"
                elif "Absence" in str(attendance.get("Status", "")):
                  attendance["Status"] = "Present(P)"

              report_actual_count += 1

          progress.progress(
              0.88,
              text="المرحلة 3/4 — مطابقة الموظفين وتجهيز القيم...",
          )

          recovered_single_punch = 0
          for attendance_date, date_map in all_attendance.items():
            for employee_id, attendance in date_map.items():
              if employee_id not in target_employee_ids:
                continue
              status = str(attendance.get("Status", "") or "")
              if "Leave" in status or "Absence" in status:
                continue
              clock_in = str(attendance.get("Clock In", "") or "").strip()
              clock_out = str(attendance.get("Clock Out", "") or "").strip()
              total_work = str(
                  attendance.get("Calculated WT", "")
                  or attendance.get("Actual WT", "")
                  or ""
              ).strip()
              if bool(clock_in) != bool(clock_out) and total_work:
                recovered_single_punch += 1

          progress.progress(
              1.0,
              text="المرحلة 4/4 — البيانات جاهزة للتصدير.",
          )
          progress.empty()
          hide_loading_overlay(monthly_loading_overlay)
          monthly_loading_overlay = None

          # EXACT MATCH ONLY:
          # Excel Column B (BioTime ID) == app/BioTime ID (emp_code).
          # No name fallback and no fuzzy matching.
          employee_matches = []
          unmatched_employees = []
          duplicate_excel_ids = set()
          seen_excel_ids = set()

          for excel_employee in excel_employees:
            excel_id = excel_employee["biotime_id"]

            if not excel_id:
              unmatched_employees.append(
                  {
                      "row": excel_employee["row"],
                      "biotime_id": "",
                      "excel_name": excel_employee["name"],
                      "reason": "BioTime ID is empty",
                  }
              )
              continue

            if excel_id in seen_excel_ids:
              duplicate_excel_ids.add(excel_id)
            seen_excel_ids.add(excel_id)

            if excel_id in employee_catalog:
              employee_matches.append(
                  {
                      "row": excel_employee["row"],
                      "excel_name": excel_employee["name"],
                      "employee_id": excel_id,
                      "api_name": employee_catalog[excel_id],
                      "score": 100,
                      "match_type": "Exact ID",
                  }
              )
            else:
              unmatched_employees.append(
                  {
                      "row": excel_employee["row"],
                      "biotime_id": excel_id,
                      "excel_name": excel_employee["name"],
                      "reason": "BioTime ID not found in app/BioTime ID list",
                  }
              )

          # Add only the BioTime employees explicitly approved by the user. They are
          # placed in reserved blank roster rows and then processed exactly like every
          # other employee. Existing named Excel rows are never overwritten.
          for approved_employee in approved_new_staff_rows:
            employee_matches.append(
                {
                    "row": approved_employee["row"],
                    "excel_name": approved_employee["name"],
                    "employee_id": approved_employee["employee_id"],
                    "api_name": approved_employee["name"],
                    "score": 100,
                    "match_type": "Approved New BioTime Staff",
                }
            )

          # Automatically generate the completed attendance file immediately after upload.
          export_loading_overlay = show_loading_overlay(
              "جاري تحديث القيم وتجهيز ملف Excel النهائي..."
          )
          filled_cells = 0
          cell_updates = []
          import_log = []

          # Write the approved employee's BioTime ID and name into the reserved blank
          # Excel row. Column A is deliberately left exactly as the template provided it.
          for approved_employee in approved_new_staff_rows:
            cell_updates.append(
                {
                    "row": approved_employee["row"],
                    "column": 2,
                    "value": excel_employee_id_value(approved_employee["employee_id"]),
                    "style_id": find_existing_cell_style_id(
                        ws_target, approved_employee["row"], 2
                    ),
                }
            )
            cell_updates.append(
                {
                    "row": approved_employee["row"],
                    "column": 3,
                    "value": approved_employee["name"],
                    "style_id": find_existing_cell_style_id(
                        ws_target, approved_employee["row"], 3
                    ),
                }
            )

          # Process every date column. Existing exact-ID staff and approved new staff
          # use the same attendance calculation and the same Excel date columns.
          for match in employee_matches:
            employee_id = match["employee_id"]
            excel_row = match["row"]

            for date_column, attendance_date in date_columns.items():
              # Future dates stay unchanged/blank.
              if attendance_date > now_syria.date():
                continue

              attendance = all_attendance.get(
                  attendance_date,
                  {},
              ).get(employee_id)

              if attendance is None:
                cell_value = "A"
                status = "Absence(A)"
                clock_in = ""
                clock_out = ""
                total_work = ""
              else:
                status = str(attendance.get("Status", ""))
                clock_in = str(attendance.get("Clock In", "") or "")
                clock_out = str(attendance.get("Clock Out", "") or "")
                total_work = str(
                    attendance.get("Calculated WT", "")
                    or attendance.get("Actual WT", "")
                    or ""
                ).strip()

                if "Leave" in status:
                  cell_value = "L"
                elif "Absence" in status:
                  cell_value = "A"
                else:
                  excel_time = time_to_excel_value(total_work)
                  cell_value = excel_time if excel_time is not None else None

              cell_updates.append(
                  {
                      "row": excel_row,
                      "column": date_column,
                      "value": cell_value,
                  }
              )
              filled_cells += 1

              import_log.append(
                  [
                      employee_id,
                      match["excel_name"],
                      match["api_name"],
                      match.get("match_type", "Exact ID"),
                      attendance_date,
                      clock_in,
                      clock_out,
                      total_work,
                      status,
                  ]
              )

          total_formula_updates = build_total_formula_updates(
              ws_target, employee_matches, date_columns, now_syria.date()
          )

          try:
            export_bytes = export_template_preserving_package(
                original_template_bytes,
                ws_target.title,
                cell_updates,
                formula_updates=total_formula_updates,
                duration_style_id=duration_style_id,
            )
          except Exception as export_error:
            raise RuntimeError(
                "تعذر إنشاء ملف Excel النهائي مع الحفاظ على بنية الملف الأصلية: "
                + str(export_error)
            ) from export_error

          strlit.session_state["monthly_attendance_export"] = export_bytes
          output_extension = ".xlsm" if keep_vba else ".xlsx"
          strlit.session_state["monthly_attendance_filename"] = (
              "Attendance_Completed_"
              f"{sorted_dates[0].strftime('%Y_%m_%d')}"
              "_to_"
              f"{sorted_dates[-1].strftime('%Y_%m_%d')}"
              f"{output_extension}"
          )
          strlit.session_state["monthly_attendance_mime"] = (
              "application/vnd.ms-excel.sheet.macroEnabled.12"
              if keep_vba
              else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
          )
          strlit.session_state["monthly_attendance_import_summary"] = {
              "matched": len(employee_matches),
              "unmatched": len(unmatched_employees),
              "filled": filled_cells,
              "dates": len(dates_list),
              "single_punch": recovered_single_punch,
              "biotime_report_rows": report_actual_count,
              "added_staff": len(approved_new_staff_rows),
              "total_cutoff": (now_syria.date() - timedelta(days=1)).strftime("%d/%m/%Y"),
          }

          hide_loading_overlay(export_loading_overlay)
          export_loading_overlay = None

          strlit.success(
              f"✅ تمت تعبئة جميع التواريخ حتى {now_syria.date().strftime('%d/%m/%Y')} "
              f"({filled_cells} خانة). تم تصدير الملف مع الحفاظ على بنية Excel الأصلية."
          )

          if "monthly_attendance_export" in strlit.session_state:
            summary = strlit.session_state.get("monthly_attendance_import_summary", {})
            if summary:
              strlit.markdown(
                  '<div class="gp-section-title">✅ ملخص معالجة الشهر</div>',
                  unsafe_allow_html=True,
              )
              render_kpi_cards([
                  ("📅", "تواريخ معالجة", summary.get("dates", 0)),
                  ("☝️", "بصمة واحدة", summary.get("single_punch", 0)),
                  ("🧮", "خانات تم تجهيزها", summary.get("filled", 0)),
                  ("📡", "قيم تقرير BioTime", summary.get("biotime_report_rows", 0)),
                  ("➕", "موظفون تمت إضافتهم", summary.get("added_staff", 0)),
              ])

            strlit.markdown(
                '<div class="gp-download-ready">📥 ملف الدوام النهائي جاهز للتحميل</div>',
                unsafe_allow_html=True,
            )
            strlit.download_button(
                label="📥 تحميل ملف الدوام النهائي",
                data=strlit.session_state["monthly_attendance_export"],
                file_name=strlit.session_state["monthly_attendance_filename"],
                mime=strlit.session_state.get(
                    "monthly_attendance_mime",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
                use_container_width=True,
                key="download_monthly_attendance",
            )

        except Exception as template_error:
          hide_loading_overlay(monthly_loading_overlay)
          hide_loading_overlay(export_loading_overlay)
          strlit.error("تعذر معالجة ملف الدوام. افتح التفاصيل الفنية عند الحاجة.")
          with strlit.expander("🛠️ التفاصيل الفنية", expanded=False):
            strlit.code(str(template_error))

  with monthly_tab:
    strlit.markdown(
        '<div class="gp-report-panel">'
        '<div class="gp-report-title">📊 التقرير الشهري المباشر</div>'
        '<div class="gp-report-subtitle">بدون رفع Excel — يتم جلب الدوام مباشرة من BioTime بنفس قواعد الحساب المعتمدة.</div>'
        '</div>',
        unsafe_allow_html=True,
    )

    month_names = {
        1: "يناير", 2: "فبراير", 3: "مارس", 4: "أبريل", 5: "مايو", 6: "يونيو",
        7: "يوليو", 8: "أغسطس", 9: "سبتمبر", 10: "أكتوبر", 11: "نوفمبر", 12: "ديسمبر",
    }
    month_col, year_col = strlit.columns(2)
    with month_col:
      report_month = strlit.selectbox(
          "الشهر",
          options=list(range(1, 13)),
          index=now_syria.month - 1,
          format_func=lambda value: f"{month_names[value]} ({value:02d})",
          key="direct_monthly_report_month",
      )
    with year_col:
      report_year_options = list(range(now_syria.year, now_syria.year - 7, -1))
      report_year = strlit.selectbox(
          "السنة",
          options=report_year_options,
          index=0,
          key="direct_monthly_report_year",
      )

    include_today_report = strlit.checkbox(
        "تضمين اليوم الحالي في التقرير",
        value=False,
        key="direct_monthly_report_include_today",
        help="افتراضياً يتم إيقاف التقرير عند أمس حتى لا تدخل ساعات يوم غير مكتمل في الإجماليات.",
    )

    report_month_start = datetime(report_year, report_month, 1).date()
    report_month_last = datetime(
        report_year,
        report_month,
        calendar.monthrange(report_year, report_month)[1],
    ).date()
    report_today = now_syria.date()
    report_cutoff = report_today if include_today_report else report_today - timedelta(days=1)
    report_month_end = min(report_month_last, report_cutoff)

    if report_month_start > report_cutoff:
      strlit.info("هذا الشهر لم يبدأ بعد ضمن تاريخ التقرير المختار.")
    else:
      strlit.caption(
          f"الفترة: {report_month_start.strftime('%d/%m/%Y')} → {report_month_end.strftime('%d/%m/%Y')}"
      )

      if strlit.button(
          "📊 إنشاء التقرير الشهري",
          use_container_width=True,
          key="generate_direct_monthly_report",
      ):
        report_overlay = show_loading_overlay("جاري إعداد التقرير الشهري من BioTime...")
        report_progress = strlit.progress(0, text="1/4 — تحميل الموظفين والبصمات...")
        try:
          requested_report_dates = []
          cursor = report_month_start
          while cursor <= report_month_end:
            requested_report_dates.append(cursor)
            cursor += timedelta(days=1)

          (
              monthly_attendance,
              monthly_employee_catalog,
              monthly_internal_id_map,
              monthly_employee_details,
          ) = load_monthly_attendance_bulk(
              report_month_start,
              report_month_end,
              requested_dates=tuple(requested_report_dates),
          )

          target_ids = set(monthly_employee_details.keys())
          expected_attendance_keys = set()
          expected_single_keys = set()
          for attendance_date, date_map in monthly_attendance.items():
            for employee_id, attendance in date_map.items():
              if employee_id not in target_ids:
                continue
              status = str(attendance.get("Status", "") or "")
              if "Leave" in status or "Absence" in status:
                continue
              key = (attendance_date, employee_id)
              expected_attendance_keys.add(key)
              ci = str(attendance.get("Clock In", "") or "").strip()
              co = str(attendance.get("Clock Out", "") or "").strip()
              if bool(ci) != bool(co):
                expected_single_keys.add(key)

          report_progress.progress(0.48, text="2/4 — مطابقة Actual WT / Total WT مع BioTime...")
          calculated_rows = fetch_biotime_calculated_range(
              report_month_start,
              report_month_end,
              monthly_internal_id_map,
              expected_single_keys,
              expected_attendance_keys,
          )

          for attendance_date, report_date_map in calculated_rows.items():
            local_map = monthly_attendance.get(attendance_date, {})
            for employee_id, report_record in report_date_map.items():
              attendance = local_map.get(employee_id)
              if attendance is None:
                continue
              actual_wt = str(report_record.get("Actual WT", "") or "").strip()
              total_wt = str(report_record.get("Report Total WT", "") or "").strip()
              report_clock_in = str(report_record.get("Clock In", "") or "").strip()
              report_clock_out = str(report_record.get("Clock Out", "") or "").strip()
              report_name = str(report_record.get("Report Name", "") or "")
              local_work = str(attendance.get("Calculated WT", "") or "").strip()

              if (report_clock_in or report_clock_out) and (
                  report_name == "monthlyWorkHoursReport" or not local_work
              ):
                attendance["Clock In"] = report_clock_in
                attendance["Clock Out"] = report_clock_out

              final_in = str(attendance.get("Clock In", "") or "").strip()
              final_out = str(attendance.get("Clock Out", "") or "").strip()
              is_single = bool(final_in) != bool(final_out)
              report_work = (actual_wt or total_wt) if is_single else (total_wt or actual_wt)

              # Never let a report endpoint replace a correct 4/6/8... odd-even sum.
              if attendance.get("Odd Even Paired") and local_work:
                selected_work = local_work
              elif report_name == "monthlyWorkHoursReport" and report_work:
                selected_work = report_work
              else:
                selected_work = local_work or report_work

              if not selected_work:
                continue
              attendance["Calculated WT"] = selected_work
              if is_single:
                attendance["Actual WT"] = selected_work
              else:
                attendance["Total WT"] = selected_work

              if "Leave" not in str(attendance.get("Status", "")):
                if final_in and not final_out:
                  attendance["Status"] = "Present(P) / Missing OUT"
                elif final_out and not final_in:
                  attendance["Status"] = "Present(P) / Missing IN"
                elif "Absence" in str(attendance.get("Status", "")):
                  attendance["Status"] = "Present(P)"

          report_progress.progress(0.73, text="3/4 — تجهيز جدول التدقيق والبصمات الخام...")
          token = get_auth_token()
          raw_report_transactions = []
          if token:
            report_session = requests.Session()
            report_headers = {
                "Authorization": f"Token {token}",
                "Content-Type": "application/json",
            }
            raw_report_transactions = fetch_paginated_api(
                report_session,
                "/iclock/api/transactions/",
                {
                    "start_time": report_month_start.strftime("%Y-%m-%d") + " 00:00:00",
                    "end_time": (report_month_end + timedelta(days=1)).strftime("%Y-%m-%d") + " 05:00:00",
                    "page_size": 5000,
                },
                report_headers,
                timeout=30,
            )

          report_progress.progress(0.90, text="4/4 — إنشاء ملف Excel المنظم...")
          report_bytes, report_summary, report_exceptions = build_monthly_attendance_workbook(
              monthly_attendance,
              monthly_employee_details,
              report_month_start,
              report_month_end,
              raw_transactions=raw_report_transactions,
          )

          report_key = f"{report_year}_{report_month:02d}_{report_month_end.strftime('%d')}"
          strlit.session_state["direct_monthly_report_bytes"] = report_bytes
          strlit.session_state["direct_monthly_report_summary"] = report_summary
          strlit.session_state["direct_monthly_report_exceptions"] = report_exceptions
          strlit.session_state["direct_monthly_report_filename"] = (
              f"BioTime_Monthly_Attendance_{report_year}_{report_month:02d}_to_{report_month_end.strftime('%d')}.xlsx"
          )
          strlit.session_state["direct_monthly_report_key"] = report_key

          report_progress.progress(1.0, text="تم تجهيز التقرير.")
          report_progress.empty()
          hide_loading_overlay(report_overlay)
          report_overlay = None
        except Exception as report_error:
          report_progress.empty()
          hide_loading_overlay(report_overlay)
          report_overlay = None
          strlit.error("تعذر إنشاء التقرير الشهري.")
          with strlit.expander("🛠️ التفاصيل الفنية", expanded=False):
            strlit.code(str(report_error))

      current_direct_report_key = f"{report_year}_{report_month:02d}_{report_month_end.strftime('%d')}"
      if (
          "direct_monthly_report_bytes" in strlit.session_state
          and strlit.session_state.get("direct_monthly_report_key") == current_direct_report_key
      ):
        summary = strlit.session_state.get("direct_monthly_report_summary", {})
        strlit.markdown(
            f'<div class="gp-report-ready">✅ التقرير جاهز — {summary.get("from", "")} إلى {summary.get("to", "")}</div>',
            unsafe_allow_html=True,
        )
        render_kpi_cards([
            ("👥", "موظفون", summary.get("employees", 0)),
            ("📅", "أيام دوام", summary.get("working_days", 0)),
            ("⏱️", "إجمالي الساعات", summary.get("total_hours", "0:00")),
            ("⏰", "تأخير", summary.get("late", 0)),
            ("❌", "غياب", summary.get("absent", 0)),
            ("☝️", "بصمة واحدة", summary.get("single_punch", 0)),
        ])

        monthly_exceptions = strlit.session_state.get("direct_monthly_report_exceptions", [])
        late_rows = [row for row in monthly_exceptions if row.get("Day Status") == "Late"]
        absent_rows = [row for row in monthly_exceptions if row.get("Day Status") == "Absent"]
        single_rows = [row for row in monthly_exceptions if row.get("Single Punch") == "Yes"]
        exc_cols = strlit.columns(3)
        with exc_cols[0]:
          if strlit.button(f"⏰ التأخير\n{len(late_rows)}", use_container_width=True, key="monthly_popup_late"):
            _popup_dataframe("⏰ التأخير الشهري", late_rows, "monthly_late")
        with exc_cols[1]:
          if strlit.button(f"❌ الغياب\n{len(absent_rows)}", use_container_width=True, key="monthly_popup_absent"):
            _popup_dataframe("❌ الغياب الشهري", absent_rows, "monthly_absent")
        with exc_cols[2]:
          if strlit.button(f"☝️ بصمة واحدة\n{len(single_rows)}", use_container_width=True, key="monthly_popup_single"):
            _popup_dataframe("☝️ حالات البصمة الواحدة", single_rows, "monthly_single")

        strlit.download_button(
            "📥 تحميل التقرير الشهري الكامل",
            data=strlit.session_state["direct_monthly_report_bytes"],
            file_name=strlit.session_state["direct_monthly_report_filename"],
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="download_direct_monthly_report",
        )

  with backup_tab:
    # 🛡️ INDEPENDENT BIOTIME BACKUP
    strlit.caption(
        "ينشئ ملف ZIP للتحميل يحتوي على Raw JSON + CSV + SQLite + الصور/الوسائط "
        "التي يسمح BioTime Cloud بقراءتها. العملية للقراءة فقط ولا تعدّل أي بيانات."
    )
    strlit.warning(
        "ملف النسخة الاحتياطية يحتوي على بيانات موظفين حساسة. احفظه في مكان خاص وآمن."
    )

    if strlit.button(
        "🛡️ إنشاء نسخة احتياطية الآن",
        use_container_width=True,
        key="create_biotime_backup",
    ):
      backup_overlay = show_loading_overlay("جاري إنشاء نسخة BioTime الاحتياطية...")
      backup_progress = strlit.progress(0, text="بدء النسخ الاحتياطي...")
      try:
        def update_backup_progress(value, message):
          backup_progress.progress(min(max(float(value), 0.0), 1.0), text=message)

        backup_bytes, backup_manifest = build_biotime_backup(update_backup_progress)
        timestamp = datetime.now(SYRIA_TZ).strftime("%Y-%m-%d_%H%M")
        strlit.session_state["biotime_backup_bytes"] = backup_bytes
        strlit.session_state["biotime_backup_filename"] = f"BioTime_Backup_{timestamp}.zip"
        strlit.session_state["biotime_backup_manifest"] = backup_manifest
        backup_progress.empty()
        hide_loading_overlay(backup_overlay)
        backup_overlay = None
      except Exception as backup_error:
        backup_progress.empty()
        hide_loading_overlay(backup_overlay)
        backup_overlay = None
        strlit.error("تعذر إنشاء النسخة الاحتياطية.")
        with strlit.expander("🛠️ تفاصيل خطأ النسخ الاحتياطي", expanded=False):
          strlit.code(str(backup_error))

    if "biotime_backup_bytes" in strlit.session_state:
      backup_manifest = strlit.session_state.get("biotime_backup_manifest", {})
      dataset_summary = backup_manifest.get("datasets", {})
      successful_sets = sum(
          1 for info in dataset_summary.values() if info.get("status_code") == 200
      )
      total_rows = sum(
          int(info.get("rows", 0) or 0) for info in dataset_summary.values()
      )
      strlit.success(
          f"النسخة جاهزة: {successful_sets} مجموعات بيانات، {total_rows:,} سجل محفوظ."
      )
      strlit.download_button(
          "📥 تحميل نسخة BioTime الاحتياطية",
          data=strlit.session_state["biotime_backup_bytes"],
          file_name=strlit.session_state["biotime_backup_filename"],
          mime="application/zip",
          use_container_width=True,
          key="download_biotime_backup",
      )
      failed_sets = [
          name for name, info in dataset_summary.items()
          if info.get("status_code") != 200
      ]
      if failed_sets:
        strlit.caption(
            "ملاحظة: بعض واجهات BioTime غير متاحة لهذا الاشتراك/الإصدار: "
            + ", ".join(failed_sets)
        )


  # Attendance details are intentionally not rendered inline below the monthly
  # section. Click one of the six summary cards above to open a clean popup.


except Exception as e:
  hide_loading_overlay(main_loading_overlay)
  strlit.error("تعذر تحميل بيانات BioTime حالياً. حاول التحديث مرة أخرى.")
  with strlit.expander("🛠️ التفاصيل الفنية", expanded=False):
    strlit.code(TEXT_CONFIG["err_api"].format(str(e)))
